"""auto_save 防抖合并保存冒烟测试。

覆盖:
1. 防抖窗口内连续写: 不逐次落盘, pending 标记正确
2. flush_save 显式落盘: pending 清空, 文件内容包含全部写操作结果
3. 读原子走内存: 防抖窗口内写的值立即可读(不依赖落盘)
4. 跨窗口写: 距上次落盘超过防抖窗口时立即落盘

运行: cd engine/components/astronverse-datatable && .venv/bin/python tests/smoke/smoke_save_debounce.py
"""

import os
import sys
import time
from unittest.mock import patch

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "src"))

import astronverse.datatable.datatable as dt_mod  # noqa: E402
from astronverse.datatable import WriteType  # noqa: E402
from astronverse.datatable.datatable import DataTable  # noqa: E402

PASS = 0
FAIL = 0


def check(name: str, cond: bool, detail: str = ""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"[PASS] {name}")
    else:
        FAIL += 1
        print(f"[FAIL] {name} {detail}")


dt = DataTable()

# 重置防抖状态, 保证测试起点确定
dt_mod._save_state["last_save"] = 0.0
dt_mod._save_state["pending"] = False

# ---------- 1. 防抖窗口内连续写不逐次落盘 ----------
save_calls = []
orig_save = dt_mod.PyxlWrapper.save


def counting_save(*a, **kw):
    save_calls.append(time.time())
    return orig_save(*a, **kw)


with patch.object(dt_mod.PyxlWrapper, "save", side_effect=counting_save):
    for i in range(5):
        dt.write_data(write_type=WriteType.CELL, row=100 + i, col="A", data=f"debounce-{i}")

check("防抖: 5次连写仅首写落盘1次", len(save_calls) == 1, f"save调用{len(save_calls)}次")
check("防抖: pending已置位", dt_mod._save_state["pending"] is True)

# ---------- 2. 读原子走内存, 防抖窗口内写的值立即可读 ----------
from astronverse.datatable import ReadType  # noqa: E402

val = dt.read_data(read_type=ReadType.CELL, row=100, col="A")
check("读内存: 未落盘的值立即可读", val == "debounce-0", f"got {val}")

# ---------- 3. flush_save 显式落盘 ----------
with patch.object(dt_mod.PyxlWrapper, "save", side_effect=counting_save):
    dt_mod.flush_save()
check("flush: 落盘1次且pending清空", len(save_calls) == 2 and dt_mod._save_state["pending"] is False)

# 落盘后重开文件验证持久化包含全部5次写(绕过内存wrapper直接读文件)
from openpyxl import load_workbook  # noqa: E402

wb = load_workbook(dt_mod._xlsx_file_path, read_only=True)
ws = wb.active
persisted = [ws.cell(row=100 + i, column=1).value for i in range(5)]
wb.close()
check("flush: 文件包含全部5次写", persisted == [f"debounce-{i}" for i in range(5)], str(persisted))

# ---------- 4. 跨窗口写立即落盘 ----------
save_calls.clear()
dt_mod._save_state["last_save"] = time.time() - dt_mod._SAVE_DEBOUNCE_SECONDS - 0.1
dt_mod._save_state["pending"] = False
with patch.object(dt_mod.PyxlWrapper, "save", side_effect=counting_save):
    dt.write_data(write_type=WriteType.CELL, row=200, col="A", data="cross-window")
check("跨窗口: 立即落盘", len(save_calls) == 1 and dt_mod._save_state["pending"] is False)

# 清理测试写入(不污染全局文件, 供后续测试/复跑)
for r in (100, 101, 102, 103, 104, 200):
    try:
        dt_mod.PyxlWrapper.sheet.cell(row=r, column=1).value = None
    except Exception:
        pass
dt_mod.flush_save()

print(f"\n结果: {PASS} 通过, {FAIL} 失败")
sys.exit(1 if FAIL else 0)
