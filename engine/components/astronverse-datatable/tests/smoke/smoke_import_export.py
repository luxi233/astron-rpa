# -*- coding: utf-8 -*-
"""P0 导入导出完整冒烟(datatable组件)

覆盖: CSV编码(utf8/gbk回退/BOM/显式ANSI)/表头导入, XLSX多表/sheet不存在报错,
XLS(xlrd)类型转换(int/float/bool/中文)/多表/报错, 密码解密错误路径+临时文件后缀,
导出CSV(覆盖/追加/编码)/JSON/XLSX/时间戳重命名, xlsx与csv往返一致性。
.xls 测试文件由内置 BIFF8 生成器产生(不引入 xlwt 依赖)。
"""

import json
import os
import struct
import sys
import tempfile
import types

sys.path.insert(0, "/Users/infinitelab/Desktop/astron-rpa/engine/components/astronverse-datatable/src")

import openpyxl as _opx

from astronverse.datatable import datatable as dt_mod
from astronverse.datatable.datatable import DataTable
from astronverse.datatable import CsvWriteType, ExportFileType, FileEncodingType, ReadType

PyxlWrapper = dt_mod.PyxlWrapper
PyxlHeadWrapper = dt_mod.PyxlHeadWrapper

passed = failed = 0


def check(name, cond, detail=""):
    global passed, failed
    if cond:
        passed += 1
        print(f"PASS {name}")
    else:
        failed += 1
        print(f"FAIL {name} {detail}")


def read_area():
    return DataTable.read_data(read_type=ReadType.AREA, start_row=1, start_col="A", end_row=-1, end_col="-1")


# ----------------------------------------------------------------------------
# BIFF8 .xls 生成器(最小记录集: BOF/BOUNDSHEET/LABEL/NUMBER/BOOLERR/EOF)
# ----------------------------------------------------------------------------
def _biff_rec(rid, data=b""):
    return struct.pack("<HH", rid, len(data)) + data


def _biff_cell(r, c, v):
    if v is None:
        return b""
    if isinstance(v, bool):
        return _biff_rec(0x0205, struct.pack("<HHHBB", r, c, 0, int(v), 0))  # BOOLERR
    if isinstance(v, (int, float)):
        return _biff_rec(0x0203, struct.pack("<HHH", r, c, 0) + struct.pack("<d", float(v)))  # NUMBER
    s = str(v)
    # LABEL: row(2) col(2) xf(2) cch(2) flags(1=UNICODE-16) chars(2*n)
    return _biff_rec(0x0204, struct.pack("<HHH", r, c, 0) + struct.pack("<H", len(s)) + b"\x01" + s.encode("utf-16-le"))


def make_xls(sheets):
    """生成多工作表 .xls(裸 BIFF8 流)。sheets: list[(sheet_name, rows_2d)]"""

    def sheet_stream(rows):
        buf = _biff_rec(0x0809, struct.pack("<HH", 0x0600, 0x0010))  # BOF worksheet
        for r, row in enumerate(rows):
            for c, v in enumerate(row):
                buf += _biff_cell(r, c, v)
        return buf + _biff_rec(0x000A)

    def boundsheet(pos, name):
        enc = name.encode("utf-16-le")
        # lbPlyPos(4) vis(1) type(1) cch(1) flags(1=UNICODE-16) chars
        return _biff_rec(0x0085, struct.pack("<iBB", pos, 0, 0) + struct.pack("<B", len(name)) + b"\x01" + enc)

    bodies = [sheet_stream(rows) for _, rows in sheets]
    names = [n for n, _ in sheets]
    globals_len = (
        len(_biff_rec(0x0809, struct.pack("<HH", 0x0600, 0x0005)))
        + sum(len(boundsheet(0, n)) for n in names)
        + len(_biff_rec(0x000A))
    )
    # 每个 sheet 的绝对偏移 = globals_len + 前面 sheet body 长度累加
    positions = []
    pos = globals_len
    for body in bodies:
        positions.append(pos)
        pos += len(body)
    out = _biff_rec(0x0809, struct.pack("<HH", 0x0600, 0x0005))  # BOF globals
    for name, p in zip(names, positions):
        out += boundsheet(p, name)
    out += _biff_rec(0x000A)  # EOF globals
    for body in bodies:
        out += body
    return out


workdir = tempfile.mkdtemp()


def _path(name):
    return os.path.join(workdir, name)


# ----------------------------------------------------------------------------
# 1. CSV 导入
# ----------------------------------------------------------------------------
csv_utf8 = _path("t_utf8.csv")
with open(csv_utf8, "w", encoding="utf-8") as f:
    f.write("名称,数量\n苹果,3\n香蕉,5\n")
DataTable.import_data_table_from_file(import_file_path=csv_utf8)
check("csv import utf8 auto", read_area() == [["名称", "数量"], ["苹果", "3"], ["香蕉", "5"]], read_area())

csv_gbk = _path("t_gbk.csv")
with open(csv_gbk, "w", encoding="gbk") as f:
    f.write("名称,数量\n苹果,3\n")
DataTable.import_data_table_from_file(import_file_path=csv_gbk)
check("csv import gbk fallback", read_area() == [["名称", "数量"], ["苹果", "3"]], read_area())

csv_bom = _path("t_bom.csv")
with open(csv_bom, "w", encoding="utf-8-sig") as f:
    f.write("a,b\n1,2\n")
DataTable.import_data_table_from_file(import_file_path=csv_bom)
check("csv import utf8-sig", read_area() == [["a", "b"], ["1", "2"]], read_area())

DataTable.import_data_table_from_file(import_file_path=csv_gbk, file_encoding=FileEncodingType.ANSI)
check("csv import ansi explicit", read_area() == [["名称", "数量"], ["苹果", "3"]], read_area())

csv_head = _path("t_head.csv")
with open(csv_head, "w", encoding="utf-8") as f:
    f.write("列A,列B\nv1,v2\nv3,v4\n")
DataTable.import_data_table_from_file(import_file_path=csv_head, first_row_is_header=True)
check(
    "csv import header data rows",
    read_area() == [["v1", "v2"], ["v3", "v4"]],
    read_area(),
)
check(
    "csv import header head row",
    [PyxlHeadWrapper.read_cell(row=1, col=i) for i in (1, 2)] == ["列A", "列B"],
    [PyxlHeadWrapper.read_cell(row=1, col=i) for i in (1, 2)],
)

# ----------------------------------------------------------------------------
# 2. XLSX 导入
# ----------------------------------------------------------------------------
xlsx_multi = _path("t_multi.xlsx")
wb = _opx.Workbook()
ws1 = wb.active
ws1.title = "第一表"
for row in [["h1", "h2"], [1, 2.5], [True, "张三"]]:
    ws1.append(row)
ws2 = wb.create_sheet("第二表")
for row in [["s2a", "s2b"], ["x", "y"]]:
    ws2.append(row)
wb.save(xlsx_multi)

DataTable.import_data_table_from_file(import_file_path=xlsx_multi)
check(
    "xlsx import default first sheet",
    read_area() == [["h1", "h2"], [1, 2.5], [True, "张三"]],
    read_area(),
)

DataTable.import_data_table_from_file(import_file_path=xlsx_multi, sheet_name="第二表")
check("xlsx import named sheet", read_area() == [["s2a", "s2b"], ["x", "y"]], read_area())

try:
    DataTable.import_data_table_from_file(import_file_path=xlsx_multi, sheet_name="不存在表")
    check("xlsx import sheet missing error", False)
except BaseException as e:
    check("xlsx import sheet missing error", "工作表不存在" in str(e), str(e)[:80])

# ----------------------------------------------------------------------------
# 3. XLS 导入(xlrd)
# ----------------------------------------------------------------------------
xls_multi = _path("t_multi.xls")
rows_s1 = [["名称", 1, 1.5, True], ["张三", 2, 2.5, False]]
rows_s2 = [["s2a", "s2b"], ["x", 9]]
with open(xls_multi, "wb") as f:
    f.write(make_xls([("第一表", rows_s1), ("第二表", rows_s2)]))

DataTable.import_data_table_from_file(import_file_path=xls_multi)
got = read_area()
check(
    "xls import type conversion",
    got == [["名称", 1, 1.5, True], ["张三", 2, 2.5, False]],
    got,
)
check("xls import int stays int", isinstance(got[0][1], int) and not isinstance(got[0][1], bool), type(got[0][1]))
check("xls import float stays float", isinstance(got[0][2], float), type(got[0][2]))
check("xls import bool stays bool", isinstance(got[0][3], bool), type(got[0][3]))

DataTable.import_data_table_from_file(import_file_path=xls_multi, sheet_name="第二表")
check("xls import named sheet", read_area() == [["s2a", "s2b"], ["x", 9]], read_area())

try:
    DataTable.import_data_table_from_file(import_file_path=xls_multi, sheet_name="不存在表")
    check("xls import sheet missing error", False)
except BaseException as e:
    check("xls import sheet missing error", "工作表不存在" in str(e), str(e)[:80])

# ----------------------------------------------------------------------------
# 4. 密码导入
# ----------------------------------------------------------------------------
try:
    DataTable.import_data_table_from_file(import_file_path=xlsx_multi, password="wrong")
    check("import wrong password error", False)
except BaseException as e:
    check("import wrong password error", "密码错误" in str(e), str(e)[:80])

# 临时文件后缀: fake msoffcrypto 验证解密临时文件保留原扩展名(.xls 修复点)
_fake = types.ModuleType("msoffcrypto")


class _FakeOfficeFile:
    def __init__(self, f):
        pass

    def load_key(self, password=None):
        pass

    def decrypt(self, out):
        out.write(b"fake")


_fake.OfficeFile = _FakeOfficeFile
_real_ms = sys.modules.get("msoffcrypto")
sys.modules["msoffcrypto"] = _fake
# fake OfficeFile 不读内容, 但函数会 open 原文件 → 先 touch 占位文件
open(_path("enc.xls"), "wb").close()
open(_path("enc.xlsx"), "wb").close()
try:
    tp = dt_mod._decrypt_excel_to_temp_file(_path("enc.xls"), "pw")
    check("decrypt temp keeps .xls suffix", tp.endswith(".xls"), tp)
    os.path.exists(tp) and os.remove(tp)
    tp2 = dt_mod._decrypt_excel_to_temp_file(_path("enc.xlsx"), "pw")
    check("decrypt temp keeps .xlsx suffix", tp2.endswith(".xlsx"), tp2)
    os.path.exists(tp2) and os.remove(tp2)
finally:
    if _real_ms is not None:
        sys.modules["msoffcrypto"] = _real_ms
    else:
        del sys.modules["msoffcrypto"]


# ----------------------------------------------------------------------------
# 5. 导入参数校验
# ----------------------------------------------------------------------------
def _err_text(e):
    # str(e)=code.message(模板部分), 友好文案在 e.message, 合并断言
    return str(e) + " " + getattr(e, "message", "")


try:
    DataTable.import_data_table_from_file(import_file_path=_path("bad.txt"))
    check("import unsupported ext error", False)
except BaseException as e:
    check("import unsupported ext error", ".xlsx" in _err_text(e), _err_text(e)[:80])

try:
    DataTable.import_data_table_from_file(import_file_path=_path("not_exist.xlsx"))
    check("import missing file error", False)
except BaseException as e:
    check("import missing file error", "文件不存在" in _err_text(e), _err_text(e)[:80])

# ----------------------------------------------------------------------------
# 6. 导出
# ----------------------------------------------------------------------------
csv_src = _path("t_export_src.csv")
with open(csv_src, "w", encoding="utf-8") as f:
    f.write("a,b\n1,2\n3,4\n")
DataTable.import_data_table_from_file(import_file_path=csv_src)

out_csv = DataTable.export_data_table_to_file(
    export_dest_path=workdir, export_file_name="exp_csv", export_file_type=ExportFileType.CSV
)
content = open(out_csv, encoding="utf-8").read()
check("export csv overwrite", content == "a,b\r\n1,2\r\n3,4\r\n" or "a,b" in content, content[:60])

DataTable.export_data_table_to_file(
    export_dest_path=workdir,
    export_file_name="exp_csv",
    export_file_type=ExportFileType.CSV,
    csv_write_type=CsvWriteType.APPEND,
)
lines = [ln for ln in open(out_csv, encoding="utf-8").read().splitlines() if ln]
check("export csv append doubles rows", len(lines) == 6, lines)

out_bom = DataTable.export_data_table_to_file(
    export_dest_path=workdir,
    export_file_name="exp_bom",
    export_file_type=ExportFileType.CSV,
    file_encoding=FileEncodingType.UTF8_BOM,
)
check("export csv utf8-sig BOM", open(out_bom, "rb").read(3) == b"\xef\xbb\xbf")

csv_cn = _path("t_cn.csv")
with open(csv_cn, "w", encoding="utf-8") as f:
    f.write("名称\n苹果\n")
DataTable.import_data_table_from_file(import_file_path=csv_cn)
out_gbk = DataTable.export_data_table_to_file(
    export_dest_path=workdir,
    export_file_name="exp_gbk",
    export_file_type=ExportFileType.CSV,
    file_encoding=FileEncodingType.ANSI,
)
check("export csv gbk encoding", open(out_gbk, encoding="gbk").read().find("苹果") >= 0)

out_json = DataTable.export_data_table_to_file(
    export_dest_path=workdir, export_file_name="exp_json", export_file_type=ExportFileType.JSON
)
check("export json", json.load(open(out_json, encoding="utf-8")) == [["名称"], ["苹果"]])

xlsx_src = _path("t_roundtrip.xlsx")
wb = _opx.Workbook()
for row in [["a", "b"], [1, 2.5], [True, "张三"]]:
    wb.active.append(row)
wb.save(xlsx_src)
DataTable.import_data_table_from_file(import_file_path=xlsx_src)
out_xlsx = DataTable.export_data_table_to_file(
    export_dest_path=workdir, export_file_name="exp_xlsx", export_file_type=ExportFileType.XLSX
)
wb2 = _opx.load_workbook(out_xlsx)
vals = [[c.value for c in row] for row in wb2.active.iter_rows()]
check("export xlsx readback", vals == [["a", "b"], [1, 2.5], [True, "张三"]], vals)

out_ts = DataTable.export_data_table_to_file(
    export_dest_path=workdir, export_file_name="ts_out", export_file_type=ExportFileType.CSV, is_overwrite=False
)
bn = os.path.basename(out_ts)
check("export timestamp rename", bn.startswith("ts_out_2") and bn.endswith(".csv"), bn)

# ----------------------------------------------------------------------------
# 7. 往返一致性
# ----------------------------------------------------------------------------
DataTable.import_data_table_from_file(import_file_path=xlsx_src)
mid = DataTable.export_data_table_to_file(
    export_dest_path=workdir, export_file_name="rt_xlsx", export_file_type=ExportFileType.XLSX
)
DataTable.import_data_table_from_file(import_file_path=mid)
check(
    "roundtrip xlsx consistency",
    read_area() == [["a", "b"], [1, 2.5], [True, "张三"]],
    read_area(),
)

csv_rt = _path("t_rt.csv")
with open(csv_rt, "w", encoding="utf-8") as f:
    f.write("a,b\n1,2\n3,4\n")
DataTable.import_data_table_from_file(import_file_path=csv_rt)
mid_csv = DataTable.export_data_table_to_file(
    export_dest_path=workdir, export_file_name="rt_csv", export_file_type=ExportFileType.CSV
)
DataTable.import_data_table_from_file(import_file_path=mid_csv)
check("roundtrip csv consistency", read_area() == [["a", "b"], ["1", "2"], ["3", "4"]], read_area())

print(f"\n=== {passed} passed, {failed} failed ===")
sys.exit(1 if failed else 0)
