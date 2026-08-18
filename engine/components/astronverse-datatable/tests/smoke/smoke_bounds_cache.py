"""openpyxl wrapper 边界缓存 + 读路径不物化单元格 冒烟测试。

覆盖:
1. max_row/max_column 缓存正确性: 写操作增量维护、delete后失效重算、switch_sheet后重算
2. 读路径不物化: read_row/read_column/read_range/read_effective_area/last_nonempty_row
   调用后 _cells 字典不增长(旧实现会为空槽创建Cell对象)
3. 语义等价: values_only 读取与旧 cell.value 读取结果一致

运行: cd engine/components/astronverse-datatable && .venv/bin/python tests/smoke/smoke_bounds_cache.py
"""

import os
import sys
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "src"))

from openpyxl import Workbook  # noqa: E402

from astronverse.datatable.openpyxl import OpenpyxlWrapper  # noqa: E402

PASS = 0
FAIL = 0


def check(name: str, cond: bool, detail: str = ""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"[PASS] {name}")
    else:
        FAIL += 1
        print(f"[FAIL] {name} {detail}")


td = tempfile.mkdtemp()
fp = os.path.join(td, "bounds.xlsx")
wb = Workbook()
ws = wb.active
for r in range(1, 6):
    for c in range(1, 4):
        ws.cell(row=r, column=c, value=f"r{r}c{c}")
wb.save(fp)
wb.close()

w = OpenpyxlWrapper(file_path=fp)

# ---------- 1. 缓存基础正确性 ----------
check("缓存: 初始max_row=5", w.get_max_row() == 5, str(w.get_max_row()))
check("缓存: 初始max_col=3", w.get_max_column() == 3, str(w.get_max_column()))
check("缓存: 命中后不重算", w.get_max_row() == 5 and w._max_row_cache == 5)

w.write_cell(row=8, col=5, value="ext")
check(
    "缓存: write_cell扩展", w.get_max_row() == 8 and w.get_max_column() == 5, f"{w.get_max_row()},{w.get_max_column()}"
)

w.append_row(["a", "b", "c", "d", "e", "f"])
check(
    "缓存: append_row递增", w.get_max_row() == 9 and w.get_max_column() == 6, f"{w.get_max_row()},{w.get_max_column()}"
)

w.write_row(row_index=11, data=[1, 2, 3, 4, 5, 6, 7])
check(
    "缓存: write_row扩展", w.get_max_row() == 11 and w.get_max_column() == 7, f"{w.get_max_row()},{w.get_max_column()}"
)

# 与 sheet 真实值对照(openpyxl自有语义)
check("缓存: 与sheet属性一致", w.get_max_row() == w.sheet.max_row and w.get_max_column() == w.sheet.max_column)

# ---------- 2. 失效重算 ----------
w.delete_rows(idx=1, amount=2)
check("失效: delete_rows后重算", w.get_max_row() == 9, f"got {w.get_max_row()}")
check("失效: 与sheet一致", w.get_max_row() == w.sheet.max_row, f"{w.get_max_row()} vs {w.sheet.max_row}")

# ---------- 3. 读路径: 密集区域内读取不额外物化 ----------
n_before = len(w.sheet._cells)
row_vals = w.read_row(1)
col_vals = w.read_column(col_index=2)
rng_vals = w.read_range("A1:B2")
area_vals = w.read_effective_area()
n_after = len(w.sheet._cells)
check("读路径: 密集区域内4种读后_cells不增长", n_before == n_after, f"{n_before}->{n_after}")
check(
    "读行: 长度=max_col且首值非空", len(row_vals) == w.get_max_column() and row_vals[0] is not None, str(row_vals[:3])
)
check("读列: 值正确", col_vals[0] is not None and len(col_vals) == w.get_max_row(), str(col_vals[:3]))
check("读区域: 2x2", len(rng_vals) == 2 and len(rng_vals[0]) == 2)
check("读全区域: 维度正确", len(area_vals) == w.get_max_row() and len(area_vals[0]) == w.get_max_column())

# last_nonempty_row 不物化(底层删除留幻影行)
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "src"))
import astronverse.datatable.datatable as dt_mod  # noqa: E402

dt_mod.PyxlWrapper = w  # 复用同一 wrapper 实例测试模块级函数
n_before = len(w.sheet._cells)
lne = dt_mod.last_nonempty_row()
n_after = len(w.sheet._cells)
check("last_nonempty: 值正确", lne == w.sheet.max_row, f"{lne} vs {w.sheet.max_row}")
check("last_nonempty: 密集区域内不物化", n_before == n_after, f"{n_before}->{n_after}")

# ---------- 4. switch_sheet 失效 ----------
w.add_sheet(title="Sheet2")
w.switch_sheet("Sheet2")
check("切表: 新表边界为1", w.get_max_row() == 1 and w.get_max_column() == 1, f"{w.get_max_row()},{w.get_max_column()}")
w.write_cell(row=3, col=2, value="s2")
check("切表: 写入后边界正确", w.get_max_row() == 3 and w.get_max_column() == 2)
w.switch_sheet(w.workbook.sheetnames[0])
check("切回: 边界恢复", w.get_max_row() == 9 and w.get_max_column() == 7, f"{w.get_max_row()},{w.get_max_column()}")

print(f"\n结果: {PASS} 通过, {FAIL} 失败")
sys.exit(1 if FAIL else 0)
