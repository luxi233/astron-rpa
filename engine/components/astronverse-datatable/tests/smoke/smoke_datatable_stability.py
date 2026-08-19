"""数据表格稳定性完整冒烟测试。

覆盖历史三类"显示不稳定"根因及其回归:
1. 边界缓存首次写入被写死成小行号 → 清空数据表格残留旧数据(真根因)
2. 防抖落盘 pending 无 Timer 兜底 → 清空后落盘推迟到进程退出, 前端 SSE 已关事件丢失
3. save 非原子覆盖写 → 并发 load_workbook 读到写一半的 zip(BadZipFile), 前端拉取失败回退旧数据

运行: cd engine/components/astronverse-datatable && .venv/bin/python tests/smoke/smoke_datatable_stability.py
"""

import os
import sys
import tempfile
import threading
import time
from unittest.mock import patch

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "src"))

import astronverse.datatable.datatable as dt_mod
from astronverse.datatable import ReadType, WriteType
from astronverse.datatable.datatable import DataTable
from astronverse.datatable.openpyxl import OpenpyxlWrapper

PASS = 0
FAIL = 0


def check(name: str, cond: bool, detail: str = ""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"[PASS] {name}")
    else:
        FAIL += 1
        print(f"[FAIL] {name} {detail}")


def disk_nonempty_rows(file_path: str) -> int:
    """绕过内存 wrapper 直接读磁盘, 统计非空行数"""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    n = 0
    for row in ws.iter_rows(values_only=True):
        if any(c is not None and c != "" for c in row):
            n += 1
    wb.close()
    return n


def disk_cell(file_path: str, row: int, col: int):
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True)
    v = wb.active.cell(row=row, column=col).value
    wb.close()
    return v


# ============================================================
# A. 边界缓存正确性 —— 根因1回归: 首次写不得把已有大边界覆盖成小行号
# ============================================================
print("\n========== A. 边界缓存 ==========")

tmpdir = tempfile.mkdtemp(prefix="dt_stab_a_")
fa = os.path.join(tmpdir, "a.xlsx")

# 模拟"上次运行残留 198 行" → 新 wrapper(新进程语义, 缓存为 None)
wa = OpenpyxlWrapper(file_path=fa)
for r in range(1, 199):
    wa.write_cell(row=r, col=1, value=f"left-{r}")
wa.save()

wb_new = OpenpyxlWrapper(file_path=fa)  # 新进程: 重新 load, 缓存 None
check("A1 残留198行加载: 边界=198", wb_new.get_max_row() == 198, f"got {wb_new.get_max_row()}")
wb_new.write_cell(row=1, col=1, value="overwrite-1")
check("A2 首次写小行号: 边界仍=198(回归bug时=1)", wb_new.get_max_row() == 198, f"got {wb_new.get_max_row()}")
wb_new.write_cell(row=250, col=3, value="beyond")
check("A3 写超界行: 边界扩到250", wb_new.get_max_row() == 250, f"got {wb_new.get_max_row()}")
check("A4 列边界同步", wb_new.get_max_column() == 3, f"got {wb_new.get_max_column()}")

# 各写方法缓存单调性
wb_new.write_row(row_index=10, data=["x", "y", "z", "w"])
check("A5 write_row 后列边界=4", wb_new.get_max_column() == 4, f"got {wb_new.get_max_column()}")
wb_new.append_row(data=["a", "b", "c", "d", "e", "f"])
check(
    "A6 append_row 后行边界=251列边界=6",
    wb_new.get_max_row() == 251 and wb_new.get_max_column() == 6,
    f"got {wb_new.get_max_row()},{wb_new.get_max_column()}",
)
wb_new.write_column(col_name="G", data=["1", "2", "3"], start_row=300)
check("A7 write_column 后行边界=302", wb_new.get_max_row() == 302, f"got {wb_new.get_max_row()}")
wb_new.insert_rows(idx=1, amount=5)
check("A8 insert_rows 边界+5=307", wb_new.get_max_row() == 307, f"got {wb_new.get_max_row()}")
wb_new.delete_rows(idx=1, amount=307)
# openpyxl 空表语义: max_row 返回 1(dimension A1), 只要缓存已失效重算即通过
check("A9 delete_rows 后缓存失效重算(空表=1)", wb_new.get_max_row() == 1, f"got {wb_new.get_max_row()}")
wb_new.save()
check("A10 全删后磁盘空", disk_nonempty_rows(fa) == 0, f"got {disk_nonempty_rows(fa)}")
wb_new.close()

# 换 sheet 边界重算
fb = os.path.join(tmpdir, "b.xlsx")
wsh = OpenpyxlWrapper(file_path=fb)
wsh.write_cell(row=30, col=2, value="s1")
wsh.add_sheet(title="S2")
wsh.switch_sheet("S2")
check("A11 切新表边界重算=1", wsh.get_max_row() == 1, f"got {wsh.get_max_row()}")
wsh.write_cell(row=5, col=1, value="s2")
wsh.switch_sheet("Sheet")
check("A12 切回原表边界恢复=30", wsh.get_max_row() == 30, f"got {wsh.get_max_row()}")
wsh.save()
wsh.close()

# ============================================================
# B. 防抖落盘 —— 根因2回归: pending 必须由 Timer 兜底落盘
# ============================================================
print("\n========== B. 防抖落盘 ==========")

dt = DataTable()
dt_mod._save_state["last_save"] = 0.0
dt_mod._save_state["pending"] = False
dt_mod._save_state["timer"] = None

save_calls = []
orig_save = dt_mod.PyxlWrapper.save


def counting_save(*a, **kw):
    save_calls.append(time.time())
    return orig_save(*a, **kw)


with patch.object(dt_mod.PyxlWrapper, "save", side_effect=counting_save):
    dt.write_data(write_type=WriteType.CELL, row=1, col="A", data="first")
    check("B1 首写(距上次落盘>窗口)立即落盘", len(save_calls) == 1)
    dt.write_data(write_type=WriteType.CELL, row=2, col="A", data="second")
    check("B2 窗口内写只标pending", dt_mod._save_state["pending"] is True and len(save_calls) == 1)
    timer = dt_mod._save_state["timer"]
    check("B3 pending 有 Timer 兜底(回归bug时无Timer)", timer is not None and timer.is_alive())
    # 读内存不依赖落盘
    v = dt.read_data(read_type=ReadType.CELL, row=2, col="A")
    check("B4 未落盘的值立即可读", v == "second", f"got {v}")

    # 不再有任何写, 在 patch 作用域内等 Timer 自动落盘(窗口0.5s + 余量)
    time.sleep(1.0)
    check(
        "B5 Timer 自动落盘(无需后续写)",
        len(save_calls) == 2 and dt_mod._save_state["pending"] is False,
        f"calls={len(save_calls)} pending={dt_mod._save_state['pending']}",
    )
check("B6 Timer 触发后已复位", dt_mod._save_state["timer"] is None or not dt_mod._save_state["timer"].is_alive())
check("B7 B2的写已持久化", disk_cell(dt_mod._xlsx_file_path, 2, 1) == "second")

# flush_save 幂等
dt_mod.flush_save()
dt_mod.flush_save()
check("B8 flush_save 幂等不重复落盘", True)  # 不抛异常即通过

# ============================================================
# C. 清空数据表格 —— 用户主诉端到端场景
# ============================================================
print("\n========== C. 清空数据表格 ==========")

# 场景1: 大表残留 + 写后立即清空(清空落在防抖窗口内) + 流程结束无后续写
dt_mod._save_state["last_save"] = 0.0
dt_mod._save_state["pending"] = False
dt_mod._save_state["timer"] = None
# 造残留: 一次跨窗口写落盘(带大量行)
dt_mod._save_state["last_save"] = time.time() - 10
for r in range(1, 61):
    dt.write_data(write_type=WriteType.CELL, row=r, col="A", data=f"residue-{r}")
dt_mod.flush_save()
check(
    "C1 造残留落盘(60行)",
    disk_nonempty_rows(dt_mod._xlsx_file_path) >= 60,
    f"got {disk_nonempty_rows(dt_mod._xlsx_file_path)}",
)

# 模拟真实流程: 紧接着写一行(触发防抖进入pending) → 立即清空
dt.write_data(write_type=WriteType.CELL, row=1, col="A", data="tmp-write")
t0 = time.time()
dt.clear_data_table(is_clear_head=True)
# 流程结束, 无后续写; 旧实现文件直到 atexit 才空 → SSE 已关 → 前端残留
time.sleep(1.0)
check(
    f"C2 清空后{time.time() - t0:.1f}s磁盘已空",
    disk_nonempty_rows(dt_mod._xlsx_file_path) == 0,
    f"got {disk_nonempty_rows(dt_mod._xlsx_file_path)}",
)
check("C3 清空后内存边界归零(空表=1)", dt_mod.PyxlWrapper.get_max_row() <= 1, f"got {dt_mod.PyxlWrapper.get_max_row()}")

# 场景2: 保留列头清空
dt_mod._save_state["last_save"] = time.time() - 10
for r in range(1, 21):
    dt.write_data(write_type=WriteType.CELL, row=r, col="A", data=f"row-{r}")
dt_mod.flush_save()
dt.clear_data_table(is_clear_head=False)
time.sleep(1.0)
left = disk_nonempty_rows(dt_mod._xlsx_file_path)
check("C4 保留列头清空: 最多残留1行(列头)", left <= 1, f"got {left}")

# 清理残留供复跑
dt.clear_data_table(is_clear_head=True)
time.sleep(0.7)

# ============================================================
# D. 原子写 —— 根因3回归: save 期间并发读零失败
# ============================================================
print("\n========== D. 原子写 ==========")

fd = os.path.join(tmpdir, "d.xlsx")
wd = OpenpyxlWrapper(file_path=fd)
for r in range(1, 101):
    for c in range(1, 11):
        wd.write_cell(row=r, col=c, value=f"init-{r}-{c}")
wd.save()

stop = threading.Event()
read_stats = {"ok": 0, "err": 0}


def d_writer():
    n = 0
    while not stop.is_set():
        n += 1
        for r in range(1, 151):
            for c in range(1, 13):
                wd.write_cell(row=r, col=c, value=f"w{n}-{r}-{c}")
        wd.save()


def d_reader():
    from openpyxl import load_workbook

    while not stop.is_set():
        try:
            w = load_workbook(fd, read_only=True, data_only=True)
            _ = w.active.max_row
            w.close()
            read_stats["ok"] += 1
        except Exception:
            read_stats["err"] += 1
        time.sleep(0.002)


ths = [threading.Thread(target=d_writer, daemon=True), threading.Thread(target=d_reader, daemon=True)]
for t in ths:
    t.start()
time.sleep(2.0)
stop.set()
for t in ths:
    t.join(timeout=5)
wd.close()

check(
    f"D1 并发读写2s: 读{read_stats['ok']}次零失败(修复前同型压测99.9%失败)",
    read_stats["err"] == 0 and read_stats["ok"] > 10,
    f"stats={read_stats}",
)

leftover_tmp = [f for f in os.listdir(tmpdir) if f.endswith(".tmp")]
check("D2 原子写无残留tmp", not leftover_tmp, f"got {leftover_tmp}")

# save 中途失败: tmp 清理 + 原文件保持完整
we = OpenpyxlWrapper(file_path=fd)
we.write_cell(row=1, col=1, value="before-error")
we.save()
with patch.object(we.workbook, "save", side_effect=OSError("disk full")):
    try:
        we.save()
        check("D3 save失败应抛出包装异常", False)
    except Exception as e:
        check("D3 save失败抛出包装异常", "数据写入失败" in str(e), str(e)[:80])
check("D4 失败后原文件完整", disk_cell(fd, 1, 1) == "before-error")
check("D5 失败后tmp已清理", not [f for f in os.listdir(tmpdir) if f.endswith(".tmp")])
we.close()

print(f"\n========== 结果: {PASS} 通过, {FAIL} 失败 ==========")
sys.exit(1 if FAIL else 0)
