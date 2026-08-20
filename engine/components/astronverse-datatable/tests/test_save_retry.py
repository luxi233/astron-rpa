import os
import tempfile
import unittest
from unittest import mock

import openpyxl

from astronverse.datatable import openpyxl as openpyxl_module
from astronverse.datatable.openpyxl import OpenpyxlWrapper


class TestSaveReplaceRetry(unittest.TestCase):
    """save() 原子替换(os.replace)撞上并发读句柄时的短重试行为。

    Windows 下 Python 的 open 不带 FILE_SHARE_DELETE, scheduler/前端并发
    load_workbook(read_only) 的瞬时读句柄会让 os.replace 抛 PermissionError:
    - 瞬时锁: 重试窗口内释放 → 保存成功
    - 持久锁(用户用 Excel/WPS 打开): 重试耗尽 → 抛友好错误并清理临时文件
    """

    def setUp(self):
        self.tmpdir = os.path.realpath(tempfile.mkdtemp(prefix="dt_save_retry_"))
        self.file_path = os.path.join(self.tmpdir, "data_table.xlsx")
        wrapper = OpenpyxlWrapper(file_path=self.file_path)
        wrapper.write_cell(row=1, col=1, value="v1")
        wrapper.save()
        wrapper.close()
        self.wrapper = OpenpyxlWrapper(file_path=self.file_path)

    def tearDown(self):
        self.wrapper.close()
        for name in os.listdir(self.tmpdir):
            os.remove(os.path.join(self.tmpdir, name))
        os.rmdir(self.tmpdir)

    def _tmp_files(self):
        return [n for n in os.listdir(self.tmpdir) if n.endswith(".tmp")]

    def test_replace_transient_lock_recovered(self):
        """前 3 次 replace 报 PermissionError(模拟瞬时读句柄), 第 4 次成功"""
        real_replace = os.replace
        calls = {"n": 0}

        def flaky_replace(src, dst):
            calls["n"] += 1
            if calls["n"] <= 3:
                raise PermissionError(5, "Access is denied")
            return real_replace(src, dst)

        self.wrapper.write_cell(row=2, col=1, value="v2")
        with (
            mock.patch.object(openpyxl_module.os, "replace", side_effect=flaky_replace),
            mock.patch.object(openpyxl_module.time, "sleep") as sleep_mock,
        ):
            self.wrapper.save()

        self.assertEqual(calls["n"], 4)
        self.assertEqual(sleep_mock.call_count, 3)
        self.assertEqual(self._tmp_files(), [])

        # 磁盘内容确已更新
        wb = openpyxl.load_workbook(self.file_path)
        self.assertEqual(wb.active.cell(row=2, column=1).value, "v2")
        wb.close()

    def test_replace_persistent_lock_raises_friendly_error(self):
        """持久锁: 重试耗尽后抛「写入数据权限被拒绝」并清理临时文件"""
        calls = {"n": 0}

        def locked_replace(src, dst):
            calls["n"] += 1
            raise PermissionError(5, "Access is denied")

        self.wrapper.write_cell(row=3, col=1, value="v3")
        with (
            mock.patch.object(openpyxl_module.os, "replace", side_effect=locked_replace),
            mock.patch.object(openpyxl_module.time, "sleep"),
        ):
            with self.assertRaises(Exception) as ctx:
                self.wrapper.save()

        self.assertEqual(calls["n"], openpyxl_module.OpenpyxlWrapper._REPLACE_RETRY_TIMES)
        self.assertIn("写入数据权限被拒绝", str(ctx.exception))
        self.assertEqual(self._tmp_files(), [])

    def test_replace_without_conflict_single_attempt(self):
        """无冲突时首次 replace 即成功, 不引入额外重试/延迟"""
        sleep_mock = mock.Mock()
        with mock.patch.object(openpyxl_module.time, "sleep", sleep_mock):
            self.wrapper.write_cell(row=4, col=1, value="v4")
            self.wrapper.save()

        sleep_mock.assert_not_called()
        self.assertEqual(self._tmp_files(), [])
        wb = openpyxl.load_workbook(self.file_path)
        self.assertEqual(wb.active.cell(row=4, column=1).value, "v4")
        wb.close()


if __name__ == "__main__":
    unittest.main()
