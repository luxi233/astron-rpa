import ast
import atexit
import csv
import json
import os
import re
import shutil
import subprocess
import sys
import time
from datetime import datetime
from functools import wraps

from astronverse.actionlib import DynamicsItem
from astronverse.actionlib.atomic import AtomicFormType, AtomicFormTypeMeta, atomicMg
from astronverse.baseline.logger.logger import logger
from astronverse.datatable import (
    AppendShift,
    BaseOperateType,
    BorderStyleType,
    CellInsertShift,
    ColumnInfoGetType,
    ColumnInsertShift,
    ConditionType,
    CopyType,
    CsvWriteType,
    DeleteCellMove,
    DeleteType,
    ExportFileType,
    FileEncodingType,
    FilterType,
    FindType,
    HAlignType,
    HideTargetType,
    InsertType,
    LoopType,
    PasteType,
    ReadType,
    RowInsertShift,
    SortOrder,
    UnderlineType,
    VAlignType,
    ValidateOperator,
    ValidateType,
    WriteMode,
    WriteType,
)
from astronverse.datatable.error import *
from astronverse.datatable.openpyxl import OpenpyxlWrapper
from astronverse.datatable.utils import (
    col_to_index,
    ensure_xlsx_file,
    filter_data,
    index_to_col,
    is_batch_spec,
    parse_col_numbers,
    parse_row_numbers,
    resolve_negative_col,
    resolve_negative_row,
    validate,
    validate_col,
    validate_end_col,
    validate_end_row,
    validate_formula,
    validate_row,
    validate_row_param,
    normalize_end_row,
    normalize_end_col,
)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

try:
    _xlsx_file_path = os.path.abspath(os.path.join(sys.exec_prefix, "../astron/data_table.xlsx"))
    _head_file_path = os.path.abspath(os.path.join(sys.exec_prefix, "../astron/data_table_head.xlsx"))
    logger.info(f"DataTable xlsx file path: {_xlsx_file_path}")
    ensure_xlsx_file(_xlsx_file_path)
    ensure_xlsx_file(_head_file_path)

    PyxlWrapper = OpenpyxlWrapper(file_path=_xlsx_file_path, sheet_name=None)
    PyxlHeadWrapper = OpenpyxlWrapper(file_path=_head_file_path, sheet_name=None)
except Exception as e:
    # 初始化失败必须留痕, 否则后续所有数据表格操作都会莫名失败且无线索
    logger.exception(f"DataTable 初始化失败, 后续数据表格操作将不可用: {e}")


# 保存防抖: 写原子高频调用时(循环写行)合并落盘, 避免每次写都全量序列化工作簿;
# 内存 wrapper 始终是最新的(读原子不受影响), 仅磁盘持久化被合并
_SAVE_DEBOUNCE_SECONDS = 0.5
_save_state = {"last_save": 0.0, "pending": False}


def flush_save():
    """立即落盘待保存的变更(进程退出前 atexit 兜底调用, 测试/外部亦可显式调用)"""
    if _save_state["pending"]:
        PyxlWrapper.save(path=_xlsx_file_path)
        _save_state["pending"] = False
    _save_state["last_save"] = time.time()


atexit.register(flush_save)


def auto_save(func):
    """自动保存装饰器(带防抖: 距上次落盘不足0.5s的写操作合并, 由下次写或退出时落盘)"""

    @wraps(func)
    def wrapper(*args, **kwargs):
        result = func(*args, **kwargs)  # type: ignore , 先执行写入操作
        if time.time() - _save_state["last_save"] >= _SAVE_DEBOUNCE_SECONDS:
            # 距上次落盘超过防抖窗口: 本次写直接落盘(刚执行过写入, 不能复用pending判断)
            PyxlWrapper.save(path=_xlsx_file_path)
            _save_state["pending"] = False
            _save_state["last_save"] = time.time()
        else:
            _save_state["pending"] = True
        return result

    return wrapper


def last_nonempty_row() -> int:
    """最后一个非空行号(openpyxl delete_rows后max_row可能残留幻影空行, 不收缩)"""
    sheet = PyxlWrapper.sheet
    max_col = PyxlWrapper.get_max_column()  # 显式传边界, 避免iter_rows默认值内部再扫一次max_column
    for r in range(PyxlWrapper.get_max_row(), 0, -1):
        # values_only 批量取值, 消除逐格属性访问开销
        row = next(sheet.iter_rows(min_row=r, max_row=r, min_col=1, max_col=max_col, values_only=True))
        if any(value is not None for value in row):
            return r
    return 0


def validate_cell(func):
    """验证行列装饰器(负数行/列号自动转换: -1表示最后一行/列)"""

    @wraps(func)
    def wrapper(*args, **kwargs):
        # 负数行号就地转换(-1=最后一行), 批量语法('1,3,5:7')与空值/0跳过
        for key in ("row", "start_row", "end_row"):
            value = kwargs.get(key)
            if value is None or value in (0, "", "0"):
                continue
            if is_batch_spec(value):
                continue
            try:
                kwargs[key] = resolve_negative_row(value)
            except (ValueError, TypeError):
                pass
        # 负数列号就地转换(-1=最后一列), 支持数字列号与字母列标, 数字统一转回字母列标
        for key in ("col", "start_col", "end_col"):
            value = kwargs.get(key)
            if value is None or value in (0, "", "0"):
                continue
            if is_batch_spec(value):
                continue
            try:
                converted = resolve_negative_col(value)
                if isinstance(converted, int):
                    converted = index_to_col(converted - 1)
                kwargs[key] = converted
            except (ValueError, TypeError):
                pass

        col = kwargs.get("col")
        row = kwargs.get("row")
        start_col = kwargs.get("start_col")
        start_row = kwargs.get("start_row")

        cols_to_validate = [c for c in [col, start_col] if c and not is_batch_spec(c)]
        for c in cols_to_validate:
            validate(col=c)
        rows_to_validate = [r for r in [row, start_row] if r and not is_batch_spec(r)]
        for r in rows_to_validate:
            validate(row=r)

        return func(*args, **kwargs)  # type: ignore

    return wrapper


def sync_data_table_head():
    """在数据表格删除列同步数据表格头部文件"""
    PyxlHeadWrapper.save(path=_head_file_path)


def _to_enum(value, enum_cls):
    """将参数规范化为枚举成员(兼容直接传枚举值字符串)"""
    if isinstance(value, enum_cls):
        return value
    return enum_cls(value)


def _resolve_format_area(format_type, row, col, start_row, start_col, end_row, end_col):
    """解析格式操作目标区域, 返回(最小行, 最大行, 最小列, 最大列)"""
    format_type = _to_enum(format_type, BaseOperateType)
    if format_type == BaseOperateType.CELL:
        if not row or not col:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("单元格操作需要指定行列"), "单元格操作需要指定行列")
        row = int(row)
        col_index = col_to_index(col)
        return row, row, col_index, col_index
    if format_type == BaseOperateType.ROW:
        if not row:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("行操作需要指定行号"), "行操作需要指定行号")
        row = int(row)
        return row, row, 1, PyxlWrapper.get_max_column()
    if format_type == BaseOperateType.COLUMN:
        if not col:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("列操作需要指定列号"), "列操作需要指定列号")
        col_index = col_to_index(col)
        return 1, PyxlWrapper.get_max_row(), col_index, col_index
    # 区域操作
    if not start_row or not start_col:
        raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("区域操作需要指定开始行列"), "区域操作需要指定开始行列")
    # 结束行/列统一归一: -1=最后(默认), 0/""=已使用区域(兼容旧语义), 负数=倒数
    end_row = normalize_end_row(end_row)
    end_col = normalize_end_col(end_col)
    start_col_index = col_to_index(start_col)
    end_col_index = col_to_index(end_col)
    start_row = int(start_row)
    end_row = int(end_row)
    if end_row < start_row:
        raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("结束行号不能小于开始行号"), "结束行号不能小于开始行号")
    if end_col_index < start_col_index:
        raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("结束列号不能小于开始列号"), "结束列号不能小于开始列号")
    return start_row, end_row, start_col_index, end_col_index


def _decrypt_excel_to_temp_file(file_path: str, password: str) -> str:
    """解密加密的Excel文件到临时文件, 返回临时文件路径"""
    import io
    import tempfile

    import msoffcrypto

    try:
        with open(file_path, "rb") as f:
            office_file = msoffcrypto.OfficeFile(f)
            office_file.load_key(password=password)
            data = io.BytesIO()
            office_file.decrypt(data)
    except DATAFRAME_EXPECTION:
        raise
    except Exception:
        raise DATAFRAME_EXPECTION(
            PARAMS_ERROR.format("密码错误或文件未加密"),
            "密码错误或文件未加密",
        )
    fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
    try:
        with os.fdopen(fd, "wb") as tmp_file:
            tmp_file.write(data.getvalue())
    except Exception:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise
    return temp_path


class DataTable:
    """数据表格"""

    @staticmethod
    @validate_cell
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param(
                "row",
                dynamics=[
                    DynamicsItem(
                        key="$this.row.show",
                        expression=f"return ['{ReadType.ROW.value}', '{ReadType.CELL.value}'].includes($this.read_type.value)",
                    )
                ],
            ),
            atomicMg.param(
                "col",
                dynamics=[
                    DynamicsItem(
                        key="$this.col.show",
                        expression=f"return ['{ReadType.COLUMN.value}', '{ReadType.CELL.value}'].includes($this.read_type.value)",
                    )
                ],
            ),
            atomicMg.param(
                "start_row",
                dynamics=[
                    DynamicsItem(
                        key="$this.start_row.show",
                        expression=f"return $this.read_type.value == '{ReadType.AREA.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "start_col",
                dynamics=[
                    DynamicsItem(
                        key="$this.start_col.show",
                        expression=f"return $this.read_type.value == '{ReadType.AREA.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "end_row",
                dynamics=[
                    DynamicsItem(
                        key="$this.end_row.show",
                        expression=f"return $this.read_type.value == '{ReadType.AREA.value}'",
                    )
                ],
                required=False,
            ),
            atomicMg.param(
                "end_col",
                dynamics=[
                    DynamicsItem(
                        key="$this.end_col.show",
                        expression=f"return $this.read_type.value == '{ReadType.AREA.value}'",
                    )
                ],
                required=False,
            ),
        ],
        outputList=[
            atomicMg.param(
                "cell_info",
                types="Str",
                dynamics=[
                    DynamicsItem(
                        key="$this.cell_info.show",
                        expression=f"return $this.read_type.value == '{ReadType.CELL.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "row_info",
                types="List",
                dynamics=[
                    DynamicsItem(
                        key="$this.row_info.show",
                        expression=f"return $this.read_type.value == '{ReadType.ROW.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "column_info",
                types="List",
                dynamics=[
                    DynamicsItem(
                        key="$this.column_info.show",
                        expression=f"return $this.read_type.value == '{ReadType.COLUMN.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "area_info",
                types="List",
                dynamics=[
                    DynamicsItem(
                        key="$this.area_info.show",
                        expression=f"return $this.read_type.value == '{ReadType.AREA.value}'",
                    )
                ],
            ),
        ],
    )
    def read_data(
        read_type: ReadType = ReadType.CELL,
        row: int = 1,
        col: str = "A",
        start_row: int = 1,
        start_col: str = "A",
        end_row: int = -1,
        end_col: str = "-1",
        is_trim_spaces: bool = False,
        is_replace_none: bool = False,
    ):
        """
        读取数据表格内容
        """
        if read_type == ReadType.CELL:
            row = validate_row_param(row)
            if not col:
                raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("读取单元格需要指定列号"), "读取单元格需要指定列号")
            col_index = col_to_index(col)
            value = PyxlWrapper.read_cell(row=row, col=col_index)
            if is_trim_spaces and isinstance(value, str):
                value = value.strip()
            if is_replace_none and value is None:
                value = ""
            return value

        if read_type == ReadType.ROW:
            row = validate_row_param(row)
            row_value = PyxlWrapper.read_row(row_index=row)
            if is_trim_spaces:
                row_value = [cell.strip() if isinstance(cell, str) else cell for cell in row_value]
            if is_replace_none:
                row_value = [cell if cell is not None else "" for cell in row_value]
            return row_value

        if read_type == ReadType.COLUMN:
            if not col:
                raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("读取列需要指定列号"), "读取列需要指定列号")
            col_index = col_to_index(col)
            col_value = PyxlWrapper.read_column(col_index=col_index)
            if is_trim_spaces:
                col_value = [cell.strip() if isinstance(cell, str) else cell for cell in col_value]
            if is_replace_none:
                col_value = [cell if cell is not None else "" for cell in col_value]
            return col_value

        if read_type == ReadType.AREA:
            start_row = validate_row_param(start_row, "开始行号")
            if not start_col:
                raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("读取区域需要指定开始列号"), "读取区域需要指定开始列号")
            # 结束行/列统一归一: -1=最后(默认), 0/""=已使用区域(兼容旧语义), 负数=倒数
            end_row = normalize_end_row(end_row)
            end_col = normalize_end_col(end_col)
            validate_col(col=end_col)
            validate_row(row=end_row)
            validate_end_col(start_col=start_col, end_col=end_col)
            validate_end_row(start_row=start_row, end_row=end_row)
            col_range = f"{start_col}{start_row}:{end_col}{end_row}"
            range_value = PyxlWrapper.read_range(range_str=col_range)
            if is_trim_spaces:
                range_value = [
                    [cell.strip() if isinstance(cell, str) else cell for cell in row_data] for row_data in range_value
                ]
            if is_replace_none:
                range_value = [[cell if cell is not None else "" for cell in row_data] for row_data in range_value]
            return range_value

    @staticmethod
    @validate_cell
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param(
                "row",
                dynamics=[
                    DynamicsItem(
                        key="$this.row.show",
                        expression=f"return ['{WriteType.ROW.value}', '{WriteType.CELL.value}'].includes($this.write_type.value)",
                    )
                ],
            ),
            atomicMg.param(
                "col",
                dynamics=[
                    DynamicsItem(
                        key="$this.col.show",
                        expression=f"return ['{WriteType.COLUMN.value}', '{WriteType.CELL.value}'].includes($this.write_type.value)",
                    )
                ],
            ),
            atomicMg.param(
                "start_row",
                dynamics=[
                    DynamicsItem(
                        key="$this.start_row.show",
                        expression=f"return ['{WriteType.AREA.value}', '{WriteType.COLUMN.value}'].includes($this.write_type.value) && $this.write_mode.value != '{WriteMode.APPEND.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "start_col",
                dynamics=[
                    DynamicsItem(
                        key="$this.start_col.show",
                        expression=f"return ['{WriteType.AREA.value}', '{WriteType.ROW.value}'].includes($this.write_type.value) && $this.write_mode.value != '{WriteMode.APPEND.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "data",
                required=True,
            ),
            atomicMg.param(
                "write_mode",
                dynamics=[
                    DynamicsItem(
                        key="$this.write_mode.show",
                        expression=f"return ['{WriteType.CELL.value}', '{WriteType.ROW.value}', '{WriteType.COLUMN.value}'].includes($this.write_type.value)",
                    )
                ],
            ),
            atomicMg.param(
                "cell_insert_shift",
                dynamics=[
                    DynamicsItem(
                        key="$this.cell_insert_shift.show",
                        expression=f"return $this.write_mode.value == '{WriteMode.INSERT.value}' && $this.write_type.value == '{WriteType.CELL.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "row_insert_shift",
                dynamics=[
                    DynamicsItem(
                        key="$this.row_insert_shift.show",
                        expression=f"return $this.write_mode.value == '{WriteMode.INSERT.value}' && $this.write_type.value == '{WriteType.ROW.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "column_insert_shift",
                dynamics=[
                    DynamicsItem(
                        key="$this.column_insert_shift.show",
                        expression=f"return $this.write_mode.value == '{WriteMode.INSERT.value}' && $this.write_type.value == '{WriteType.COLUMN.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "append_position",
                dynamics=[
                    DynamicsItem(
                        key="$this.append_position.show",
                        expression=f"return $this.write_mode.value == '{WriteMode.APPEND.value}' && $this.write_type.value == '{WriteType.CELL.value}'",
                    )
                ],
            ),
        ],
        outputList=[],
    )
    def write_data(
        write_type: WriteType = WriteType.CELL,
        row: int = 1,
        col: str = "A",
        data=None,
        start_row: int = 1,
        start_col: str = "A",
        write_mode: WriteMode = WriteMode.OVERWRITE,
        cell_insert_shift: CellInsertShift = CellInsertShift.DOWN,
        row_insert_shift: RowInsertShift = RowInsertShift.DOWN,
        column_insert_shift: ColumnInsertShift = ColumnInsertShift.RIGHT,
        append_position: AppendShift = AppendShift.ROW,
    ):
        """
        向表格写入指定数据
        """
        if data is None:
            raise DATAFRAME_EXPECTION(DATAFRAME_ERROR.format("数据不能为空"), "写入数据不能为空")

        if isinstance(data, str):
            try:
                # 尝试将字符串解析为 Python 字面量 (例如列表)
                evaluated_data = ast.literal_eval(data)
                data = evaluated_data
            except (ValueError, SyntaxError):
                pass

        if write_type == WriteType.CELL:
            row = validate_row_param(row)
            if not col:
                raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("写入单元格需要指定列号"), "写入单元格需要指定列号")
            col_index = col_to_index(col)
            if not isinstance(data, str):
                data = str(data)
            if write_mode == WriteMode.OVERWRITE:
                PyxlWrapper.write_cell(row=row, col=col_index, value=data)
            elif write_mode == WriteMode.APPEND:
                if append_position == AppendShift.ROW:
                    row_list = PyxlWrapper.read_row(row_index=row)
                    row_list.append(data)
                    PyxlWrapper.write_row(row_index=row, data=row_list)
                if append_position == AppendShift.COLUMN:
                    col_list = PyxlWrapper.read_column(col_index=col_index)
                    col_list.append(data)
                    PyxlWrapper.write_column(col_index=col_index, data=col_list)
            elif write_mode == WriteMode.INSERT:
                if cell_insert_shift == CellInsertShift.DOWN:
                    col_back = PyxlWrapper.read_column(col_index=col_index)[row - 1 :]
                    col_front = PyxlWrapper.read_column(col_index=col_index)[: row - 1]
                    new_col = col_front + [data] + col_back
                    PyxlWrapper.write_column(col_index=col_index, data=new_col)
                if cell_insert_shift == CellInsertShift.RIGHT:
                    row_back = PyxlWrapper.read_row(row_index=row)[col_index - 1 :]
                    row_front = PyxlWrapper.read_row(row_index=row)[: col_index - 1]
                    new_row = row_front + [data] + row_back
                    PyxlWrapper.write_row(row_index=row, data=new_row)
            return
        if write_type == WriteType.ROW:
            row = validate_row_param(row)
            col_index = col_to_index(start_col)
            if not isinstance(data, list):
                data = [data]
            if write_mode == WriteMode.OVERWRITE:
                PyxlWrapper.write_row(row_index=row, data=data, start_col=col_index)
            elif write_mode == WriteMode.APPEND:
                # 行追加即按行追加
                existing_row = PyxlWrapper.read_row(row_index=row)
                new_row = existing_row + data
                PyxlWrapper.write_row(row_index=row, data=new_row)
            elif write_mode == WriteMode.INSERT:
                if row_insert_shift == RowInsertShift.DOWN:
                    row_index = row + 1
                    PyxlWrapper.insert_rows(idx=row_index, amount=1)
                    PyxlWrapper.write_row(row_index=row_index, data=data, start_col=col_index)
                if row_insert_shift == RowInsertShift.UP:
                    if row == 1:
                        PyxlWrapper.insert_rows(idx=1, amount=1)
                        PyxlWrapper.write_row(row_index=1, data=data, start_col=col_index)
                    else:
                        row_index = row - 1
                        PyxlWrapper.insert_rows(idx=row_index, amount=1)
                        PyxlWrapper.write_row(row_index=row_index, data=data, start_col=col_index)

        if write_type == WriteType.COLUMN:
            if not col:
                raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("列号不能为空"), "列号不能为空")
            col_index = col_to_index(col)
            if not isinstance(data, list):
                data = [data]
            if write_mode == WriteMode.OVERWRITE:
                PyxlWrapper.write_column(col_index=col_index, data=data, start_row=start_row)
            elif write_mode == WriteMode.APPEND:
                # 列追加即按列追加
                existing_col = PyxlWrapper.read_column(col_index=col_index)
                new_col = existing_col + data
                PyxlWrapper.write_column(col_index=col_index, data=new_col)
            elif write_mode == WriteMode.INSERT:
                if column_insert_shift == ColumnInsertShift.LEFT:
                    col_index_new = col_index
                    PyxlWrapper.insert_cols(idx=col_index_new, amount=1)
                    PyxlWrapper.write_column(col_index=col_index_new, data=data, start_row=start_row)
                    PyxlHeadWrapper.insert_cols(idx=col_index_new, amount=1)
                    sync_data_table_head()
                if column_insert_shift == ColumnInsertShift.RIGHT:
                    col_index_new = col_index + 1
                    PyxlWrapper.insert_cols(idx=col_index_new, amount=1)
                    PyxlWrapper.write_column(col_index=col_index_new, data=data, start_row=start_row)
                    PyxlHeadWrapper.insert_cols(idx=col_index_new, amount=1)
                    sync_data_table_head()

        if write_type == WriteType.AREA:
            start_row = validate_row_param(start_row, "开始行号")
            if not start_col:
                raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("区域写入需要指定开始列号"), "区域写入需要指定开始列号")
            if not isinstance(data, list):
                try:
                    # 尝试将字符串解析为列表
                    data = ast.literal_eval(data)
                except Exception:
                    data = [[data]]
            else:
                data = [row_data if isinstance(row_data, list) else [row_data] for row_data in data]

            start_col_index = col_to_index(start_col)
            for i, row_data in enumerate(data):
                PyxlWrapper.write_row(row_index=start_row + i, data=row_data, start_col=start_col_index)

    @staticmethod
    @atomicMg.atomic(
        "DataTable",
        inputList=[],
        outputList=[
            atomicMg.param("max_row", types="Int"),
        ],
    )
    def get_max_row() -> int:
        """
        获取数据表格最大行号
        """
        return PyxlWrapper.get_max_row()

    @staticmethod
    @atomicMg.atomic(
        "DataTable",
        inputList=[],
        outputList=[
            atomicMg.param("max_column", types="Int"),
        ],
    )
    def get_max_column() -> int:
        """
        获取数据表格最大列号
        """
        return PyxlWrapper.get_max_column()

    @staticmethod
    @validate_cell
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param(
                "row",
                dynamics=[
                    DynamicsItem(
                        key="$this.row.show",
                        expression=f"return ['{CopyType.ROW.value}', '{CopyType.CELL.value}'].includes($this.copy_type.value)",
                    )
                ],
            ),
            atomicMg.param(
                "col",
                dynamics=[
                    DynamicsItem(
                        key="$this.col.show",
                        expression=f"return ['{CopyType.COLUMN.value}', '{CopyType.CELL.value}'].includes($this.copy_type.value)",
                    )
                ],
            ),
            atomicMg.param(
                "start_row",
                dynamics=[
                    DynamicsItem(
                        key="$this.start_row.show",
                        expression=f"return $this.copy_type.value == '{CopyType.AREA.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "start_col",
                dynamics=[
                    DynamicsItem(
                        key="$this.start_col.show",
                        expression=f"return $this.copy_type.value == '{CopyType.AREA.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "end_row",
                dynamics=[
                    DynamicsItem(
                        key="$this.end_row.show",
                        expression=f"return $this.copy_type.value == '{CopyType.AREA.value}'",
                    )
                ],
                required=False,
            ),
            atomicMg.param(
                "end_col",
                dynamics=[
                    DynamicsItem(
                        key="$this.end_col.show",
                        expression=f"return $this.copy_type.value == '{CopyType.AREA.value}'",
                    )
                ],
                required=False,
            ),
        ],
        outputList=[
            atomicMg.param(
                "copied_cell",
                types="Str",
                dynamics=[
                    DynamicsItem(
                        key="$this.copied_cell.show",
                        expression=f"return $this.copy_type.value == '{CopyType.CELL.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "copied_row",
                types="List",
                dynamics=[
                    DynamicsItem(
                        key="$this.copied_row.show",
                        expression=f"return $this.copy_type.value == '{CopyType.ROW.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "copied_column",
                types="List",
                dynamics=[
                    DynamicsItem(
                        key="$this.copied_column.show",
                        expression=f"return $this.copy_type.value == '{CopyType.COLUMN.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "copied_area",
                types="List",
                dynamics=[
                    DynamicsItem(
                        key="$this.copied_area.show",
                        expression=f"return $this.copy_type.value == '{CopyType.AREA.value}'",
                    )
                ],
            ),
        ],
    )
    def copy_data(
        copy_type: CopyType = CopyType.CELL,
        row: int = 1,
        col: str = "A",
        start_row: int = 1,
        start_col: str = "A",
        end_row: int = -1,
        end_col: str = "-1",
    ):
        """
        复制数据，复制指定单元格，行，列，区域的内容
        """

        if copy_type == CopyType.CELL and (not row or not col):
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("复制单元格需要指定行列"), "复制单元格需要指定行列")
        if copy_type == CopyType.ROW and not row:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("复制行需要指定行号"), "复制行需要指定行号")
        if copy_type == CopyType.COLUMN and not col:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("复制列需要指定列号"), "复制列需要指定列号")
        if copy_type == CopyType.AREA:
            if not start_row or not start_col:
                raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("复制区域需要指定开始行列"), "复制区域需要指定开始行列")

        # 写入到系统剪切板
        import pyperclip

        _clipboard = DataTable.read_data(
            read_type=ReadType(copy_type.value),
            row=row,
            col=col,
            start_row=start_row,
            start_col=start_col,
            end_row=end_row,
            end_col=end_col,
        )

        pyperclip.copy(str(_clipboard))
        return _clipboard

    @staticmethod
    @validate_cell
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param(
                "row",
                dynamics=[
                    DynamicsItem(
                        key="$this.row.show",
                        expression=f"return ['{PasteType.CELL.value}', '{PasteType.ROW.value}'].includes($this.paste_type.value)",
                    )
                ],
            ),
            atomicMg.param(
                "col",
                dynamics=[
                    DynamicsItem(
                        key="$this.col.show",
                        expression=f"return ['{PasteType.CELL.value}', '{PasteType.COLUMN.value}'].includes($this.paste_type.value)",
                    )
                ],
            ),
            atomicMg.param(
                "start_row",
                dynamics=[
                    DynamicsItem(
                        key="$this.start_row.show",
                        expression=f"return ['{PasteType.AREA.value}', '{PasteType.COLUMN.value}'].includes($this.paste_type.value)",
                    )
                ],
            ),
            atomicMg.param(
                "start_col",
                dynamics=[
                    DynamicsItem(
                        key="$this.start_col.show",
                        expression=f"return ['{PasteType.AREA.value}', '{PasteType.ROW.value}'].includes($this.paste_type.value)",
                    )
                ],
            ),
        ],
        outputList=[],
    )
    def paste_data(
        paste_type: PasteType = PasteType.CELL,
        row: int = 1,
        col: str = "A",
        start_row: int = 1,
        start_col: str = "A",
    ):
        """
        粘贴数据，将复制的数据粘贴到指定单元格，行，列，区域
        """
        import pyperclip

        _clipboard = pyperclip.paste()

        if paste_type == PasteType.CELL and (not row or not col):
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("粘贴单元格需要指定行列"), "粘贴单元格需要指定行列")
        if paste_type == PasteType.ROW and not row:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("粘贴行需要指定行号"), "粘贴行需要指定行号")
        if paste_type == PasteType.COLUMN and not col:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("粘贴列需要指定列号"), "粘贴列需要指定列号")
        if paste_type == PasteType.AREA and (not start_row or not start_col):
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("粘贴区域需要指定开始行列"), "粘贴区域需要指定开始行列")

        if paste_type != PasteType.CELL:
            try:
                # 使用 ast.literal_eval 代替 eval
                _clipboard = ast.literal_eval(_clipboard)
            except (ValueError, SyntaxError):
                pass

        DataTable.write_data(
            write_type=WriteType(paste_type.value),
            row=row,
            col=col,
            data=_clipboard,
            start_row=start_row,
            start_col=start_col,
            write_mode=WriteMode.OVERWRITE,  # 粘贴时均为覆盖写入
        )

    @staticmethod
    @validate_cell
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param(
                "row",
                types="Str",
                dynamics=[
                    DynamicsItem(
                        key="$this.row.show",
                        expression=f"return ['{DeleteType.CELL.value}', '{DeleteType.ROW.value}'].includes($this.delete_type.value)",
                    )
                ],
            ),
            atomicMg.param(
                "col",
                types="Str",
                dynamics=[
                    DynamicsItem(
                        key="$this.col.show",
                        expression=f"return ['{DeleteType.CELL.value}', '{DeleteType.COLUMN.value}'].includes($this.delete_type.value)",
                    )
                ],
            ),
            atomicMg.param(
                "start_row",
                dynamics=[
                    DynamicsItem(
                        key="$this.start_row.show",
                        expression=f"return $this.delete_type.value == '{DeleteType.AREA.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "start_col",
                dynamics=[
                    DynamicsItem(
                        key="$this.start_col.show",
                        expression=f"return $this.delete_type.value == '{DeleteType.AREA.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "end_row",
                types="int",
                dynamics=[
                    DynamicsItem(
                        key="$this.end_row.show",
                        expression=f"return $this.delete_type.value == '{DeleteType.AREA.value}'",
                    )
                ],
                required=False,
            ),
            atomicMg.param(
                "end_col",
                types="Str",
                dynamics=[
                    DynamicsItem(
                        key="$this.end_col.show",
                        expression=f"return $this.delete_type.value == '{DeleteType.AREA.value}'",
                    )
                ],
                required=False,
            ),
            atomicMg.param(
                "delete_cell_move",
                dynamics=[
                    DynamicsItem(
                        key="$this.delete_cell_move.show",
                        expression=f"return $this.delete_type.value == '{DeleteType.CELL.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "delete_col_move",
                dynamics=[
                    DynamicsItem(
                        key="$this.delete_col_move.show",
                        expression=f"return $this.delete_type.value == '{DeleteType.COLUMN.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "delete_row_move",
                dynamics=[
                    DynamicsItem(
                        key="$this.delete_row_move.show",
                        expression=f"return $this.delete_type.value == '{DeleteType.ROW.value}'",
                    )
                ],
            ),
        ],
        outputList=[],
    )
    def delete_data(
        delete_type: DeleteType = DeleteType.CELL,
        row: int = 1,
        col: str = "A",
        start_row: int = 1,
        start_col: str = "A",
        end_row: int = -1,
        end_col: str = "-1",
        delete_cell_move: DeleteCellMove = DeleteCellMove.UP,
        delete_col_move: bool = True,
        delete_row_move: bool = True,
    ):
        """
        删除数据表格内容
        """
        if delete_type == DeleteType.CELL:
            if not row or not col:
                raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("删除单元格需要指定行列"), "删除单元格需要指定行列")
            if is_batch_spec(row) or is_batch_spec(col):
                raise DATAFRAME_EXPECTION(
                    PARAMS_ERROR.format("删除单元格不支持批量语法，请使用行/列删除"),
                    "删除单元格不支持批量语法，请使用行/列删除",
                )
            col_index = col_to_index(col)
            PyxlWrapper.delete_cell(
                row=int(row),
                col=col_index,
                move_direction=delete_cell_move.value,
            )
        if delete_type == DeleteType.ROW:
            if not row:
                raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("删除行需要指定行号"), "删除行需要指定行号")
            # 支持批量行号: '1,3,5:7'(-1表示最后一行), 从大到小删除避免行号位移
            for row_index in sorted(parse_row_numbers(row), reverse=True):
                if delete_row_move:
                    PyxlWrapper.delete_rows(idx=row_index, amount=1)
                else:
                    PyxlWrapper.empty_row(
                        row_index=row_index,
                    )
        if delete_type == DeleteType.COLUMN:
            if not col:
                raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("删除列需要指定列号"), "删除列需要指定列号")
            # 支持批量列号: 'A,C,E:G'或'1,3,5:7'(-1表示最后一列), 从大到小删除避免列号位移
            for col_index in sorted(parse_col_numbers(col), reverse=True):
                if delete_col_move:
                    PyxlWrapper.delete_cols(idx=col_index, amount=1)
                    PyxlHeadWrapper.delete_cols(idx=col_index, amount=1)
                    sync_data_table_head()
                else:
                    PyxlWrapper.empty_column(
                        col_index=col_index,
                    )
        if delete_type == DeleteType.AREA:
            start_row = validate_row_param(start_row, "开始行号")
            if not start_col:
                raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("删除区域需要指定开始列号"), "删除区域需要指定开始列号")
            # 结束行/列统一归一: -1=最后(默认), 0/""=已使用区域(兼容旧语义), 负数=倒数
            end_row = normalize_end_row(end_row)
            end_col = normalize_end_col(end_col)
            validate_col(col=end_col)
            validate_row(row=end_row)
            validate_end_col(start_col=start_col, end_col=end_col)
            validate_end_row(start_row=start_row, end_row=end_row)
            col_range = f"{start_col}{start_row}:{end_col}{end_row}"
            PyxlWrapper.clear_range(range_str=col_range)

    @staticmethod
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        noAdvanced=True,
        inputList=[
            atomicMg.param(
                "is_clear_head",
                formType=AtomicFormTypeMeta(AtomicFormType.CHECKBOX.value),
                required=False,
            ),
        ],
        outputList=[],
    )
    def clear_data_table(is_clear_head: bool = False):
        """
        清空数据表格
        :param is_clear_head: 是否同时清空列头信息(默认保留列头)
        """
        max_row = PyxlWrapper.get_max_row()
        if max_row >= 1:
            PyxlWrapper.delete_rows(idx=1, amount=max_row)
        if is_clear_head:
            head_max_col = PyxlHeadWrapper.get_max_column()
            if head_max_col >= 1:
                PyxlHeadWrapper.delete_cols(idx=1, amount=head_max_col)
                sync_data_table_head()

    @staticmethod
    @validate_cell
    @atomicMg.atomic(
        "DataTable",
        noAdvanced=True,
        inputList=[
            atomicMg.param(
                "loop_type",
                formType=AtomicFormTypeMeta(AtomicFormType.SELECT.value),
            ),
            atomicMg.param(
                "row",
                dynamics=[
                    DynamicsItem(
                        key="$this.row.show",
                        expression=f"return $this.loop_type.value == '{LoopType.ROW.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "col",
                dynamics=[
                    DynamicsItem(
                        key="$this.col.show",
                        expression=f"return $this.loop_type.value == '{LoopType.COLUMN.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "start_row",
                dynamics=[
                    DynamicsItem(
                        key="$this.start_row.show",
                        expression=f"return $this.loop_type.value == '{LoopType.AREA.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "start_col",
                dynamics=[
                    DynamicsItem(
                        key="$this.start_col.show",
                        expression=f"return $this.loop_type.value == '{LoopType.AREA.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "end_row",
                dynamics=[
                    DynamicsItem(
                        key="$this.end_row.show",
                        expression=f"return $this.loop_type.value == '{LoopType.AREA.value}'",
                    )
                ],
                required=False,
            ),
            atomicMg.param(
                "end_col",
                dynamics=[
                    DynamicsItem(
                        key="$this.end_col.show",
                        expression=f"return $this.loop_type.value == '{LoopType.AREA.value}'",
                    )
                ],
                required=False,
            ),
        ],
        outputList=[
            atomicMg.param("index", types="Int"),
            atomicMg.param("value", types="Any"),
        ],
    )
    def loop_data_table(
        loop_type: LoopType = LoopType.ROW,
        row: int = 1,
        col: str = "A",
        start_row: int = 1,
        start_col: str = "A",
        end_row: int = -1,
        end_col: str = "-1",
    ):
        """
        遍历数据表格内容
        """
        if loop_type == LoopType.ROW and not row:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("遍历行需要指定行号"), "遍历行需要指定行号")
        if loop_type == LoopType.COLUMN and not col:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("遍历列需要指定列号"), "遍历列需要指定列号")
        if loop_type == LoopType.AREA:
            if not start_row or not start_col:
                raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("遍历区域需要指定开始行列"), "遍历区域需要指定开始行列")

        list_data = DataTable.read_data(
            read_type=ReadType(loop_type.value),
            row=row,
            col=col,
            start_row=start_row,
            start_col=start_col,
            end_row=end_row,
            end_col=end_col,
        )

        if not list_data:
            list_data = []
        if not isinstance(list_data, list):
            list_data = [list_data]

        def table_generator():
            list_length = len(list_data)
            for i in range(list_length):
                yield i, list_data[i]

        return table_generator()

    @staticmethod
    @validate_cell
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param(
                "row",
                dynamics=[
                    DynamicsItem(
                        key="$this.row.show",
                        expression=f"return $this.insert_type.value == '{InsertType.ROW.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "col",
                dynamics=[
                    DynamicsItem(
                        key="$this.col.show",
                        expression=f"return $this.insert_type.value == '{InsertType.COLUMN.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "row_insert_shift",
                dynamics=[
                    DynamicsItem(
                        key="$this.row_insert_shift.show",
                        expression=f"return $this.insert_type.value == '{InsertType.ROW.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "column_insert_shift",
                dynamics=[
                    DynamicsItem(
                        key="$this.column_insert_shift.show",
                        expression=f"return $this.insert_type.value == '{InsertType.COLUMN.value}'",
                    )
                ],
            ),
        ],
        outputList=[],
    )
    def insert_row_column(
        insert_type: InsertType = InsertType.ROW,
        row: int = 1,
        col: str = "A",
        amount: int = 1,
        row_insert_shift: RowInsertShift = RowInsertShift.DOWN,
        column_insert_shift: ColumnInsertShift = ColumnInsertShift.RIGHT,
    ):
        """
        插入行或列
        """
        if not amount:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("插入数量不能为空"), "插入数量不能为空")
        if amount < 0:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("插入数量必须大于0"), "插入数量必须大于0")
        if amount == 0:
            return
        if insert_type == InsertType.ROW:
            row = validate_row_param(row)
            if row_insert_shift == RowInsertShift.UP:
                if row == 1:
                    PyxlWrapper.insert_rows(idx=1, amount=amount)
                else:
                    PyxlWrapper.insert_rows(idx=row - 1, amount=amount)
            if row_insert_shift == RowInsertShift.DOWN:
                PyxlWrapper.insert_rows(idx=row + 1, amount=amount)
        if insert_type == InsertType.COLUMN:
            if not col:
                raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("列号不能为空"), "列号不能为空")
            col_index = col_to_index(col)
            if column_insert_shift == ColumnInsertShift.LEFT:
                pass
            if column_insert_shift == ColumnInsertShift.RIGHT:
                col_index += 1
            PyxlWrapper.insert_cols(idx=col_index, amount=amount)
            PyxlHeadWrapper.insert_cols(idx=col_index, amount=amount)
            sync_data_table_head()

    @staticmethod
    @validate_cell
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        inputList=[],
        outputList=[],
    )
    def insert_formula(
        row: int = 1,
        col: str = "A",
        formula: str = "",
    ):
        """
        插入公式到指定单元格
        """
        row = validate_row_param(row)
        if not col:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("列号不能为空"), "列号不能为空")
        if not formula:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("公式不能为空"), "公式不能为空")
        validate_formula(formula)
        col_index = col_to_index(col)
        PyxlWrapper.write_cell(row=row, col=col_index, value=formula)

    @staticmethod
    @validate_cell
    @atomicMg.atomic(
        "DataTable",
        inputList=[],
        outputList=[],
    )
    def set_column_title(
        col: str = "A",
        title: str = "",
    ):
        """
        设置列信息
        """
        if not col:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("列号不能为空"), "列号不能为空")
        if not title:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("列信息不能为空"), "列信息不能为空")
        col_index = col_to_index(col)
        PyxlHeadWrapper.write_cell(row=1, col=col_index, value=title)
        sync_data_table_head()

    @staticmethod
    @validate_cell
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param(
                "get_type",
                formType=AtomicFormTypeMeta(AtomicFormType.SELECT.value),
            ),
            atomicMg.param(
                "col",
                dynamics=[
                    DynamicsItem(
                        key="$this.col.show",
                        expression=f"return $this.get_type.value == '{ColumnInfoGetType.BY_COL.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "title",
                dynamics=[
                    DynamicsItem(
                        key="$this.title.show",
                        expression=f"return $this.get_type.value == '{ColumnInfoGetType.BY_TITLE.value}'",
                    )
                ],
            ),
        ],
        outputList=[
            atomicMg.param(
                "column_title",
                types="Str",
            ),
        ],
    )
    def get_column_title(
        col: str = "A",
        title: str = "",
        get_type: ColumnInfoGetType = ColumnInfoGetType.BY_COL,
    ) -> str:
        """
        获取列信息
        :param col: 列号(A或1, 支持负数-1表示最后一列)
        :param title: 列描述(列头信息)
        :param get_type: 获取方式(根据列号获取列描述/根据列描述获取列号)
        """
        if get_type == ColumnInfoGetType.BY_TITLE:
            if not title:
                raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("列描述不能为空"), "列描述不能为空")
            head_max_col = PyxlHeadWrapper.get_max_column()
            for i in range(1, head_max_col + 1):
                value = PyxlHeadWrapper.read_cell(row=1, col=i)
                if value is not None and str(value) == str(title):
                    return index_to_col(i - 1)
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format(f"未找到列描述为[{title}]的列"),
                f"未找到列描述为[{title}]的列",
            )
        if not col:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("列号不能为空"), "列号不能为空")
        col_index = col_to_index(col)
        return str(PyxlHeadWrapper.read_cell(row=1, col=col_index))

    @staticmethod
    @auto_save
    @validate_cell
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param(
                "find_type",
                formType=AtomicFormTypeMeta(AtomicFormType.SELECT.value),
            ),
            atomicMg.param(
                "col",
                dynamics=[
                    DynamicsItem(
                        key="$this.col.show",
                        expression=f"return $this.find_type.value == '{FindType.COLUMN.value}'",
                    )
                ],
            ),
            atomicMg.param("find_value", required=True),
            atomicMg.param(
                "replace_value",
                dynamics=[
                    DynamicsItem(
                        key="$this.replace_value.show",
                        expression="return $this.is_replace.value == true",
                    )
                ],
                required=False,
            ),
        ],
        outputList=[
            atomicMg.param(
                "find_data_positions",
                types="List",
            ),
        ],
    )
    def find_and_replace(
        find_type: FindType = FindType.TABLE,
        col: str = "A",
        find_value: str = "",
        is_case_sensitive: bool = True,
        is_replace: bool = True,
        replace_value: str = "",
    ) -> list:
        """
        查找并替换数据表格中的指定内容, 返回查找到的数据位置列表[(row, col), ...]
        """
        if not find_value:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("查找内容不能为空"), "查找内容不能为空")

        find_data_positions = []
        if find_type == FindType.COLUMN:
            if not col:
                raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("列号不能为空"), "列号不能为空")
            col_index = col_to_index(col)
            column_data = PyxlWrapper.read_column(col_index=col_index)
            for r, cell_value in enumerate(column_data, start=1):
                if cell_value is not None:
                    cell_str = str(cell_value)
                    if is_case_sensitive:
                        if find_value in cell_str:
                            find_data_positions.append((r, col))
                            if is_replace:
                                new_value = cell_str.replace(find_value, replace_value)
                                PyxlWrapper.write_cell(row=r, col=col_index, value=new_value)
                    else:
                        if find_value.lower() in cell_str.lower():
                            find_data_positions.append((r, col))
                            if is_replace:
                                new_value = re.sub(re.escape(find_value), replace_value, cell_str, flags=re.IGNORECASE)
                                PyxlWrapper.write_cell(row=r, col=col_index, value=new_value)
        else:
            max_row = PyxlWrapper.get_max_row()
            max_col = PyxlWrapper.get_max_column()
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    cell_value = PyxlWrapper.read_cell(row=r, col=c)
                    if cell_value is not None:
                        cell_str = str(cell_value)
                        found = False
                        if is_case_sensitive:
                            if find_value in cell_str:
                                found = True
                        else:
                            if find_value.lower() in cell_str.lower():
                                found = True

                        if found:
                            find_data_positions.append((r, index_to_col(c - 1)))
                            if is_replace:
                                if is_case_sensitive:
                                    new_value = cell_str.replace(find_value, replace_value)
                                else:
                                    new_value = re.sub(
                                        re.escape(find_value), replace_value, cell_str, flags=re.IGNORECASE
                                    )
                                PyxlWrapper.write_cell(row=r, col=c, value=new_value)
        return find_data_positions

    @staticmethod
    @validate_cell
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param("filter_type", formType=AtomicFormTypeMeta(AtomicFormType.SELECT.value)),
            atomicMg.param(
                "row",
                dynamics=[
                    DynamicsItem(
                        key="$this.row.show",
                        expression=f"return $this.filter_type.value == '{FilterType.ROW.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "col",
                dynamics=[
                    DynamicsItem(
                        key="$this.col.show",
                        expression=f"return $this.filter_type.value == '{FilterType.COLUMN.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "condition_value",
                dynamics=[
                    DynamicsItem(
                        key="$this.condition_value.show",
                        expression=f"return !['{ConditionType.DATE_AFTER.value}', '{ConditionType.DATE_BEFORE.value}', '{ConditionType.DATE_BETWEEN.value}'].includes($this.condition_type.value)",
                    )
                ],
            ),
            atomicMg.param(
                "date_value",
                dynamics=[
                    DynamicsItem(
                        key="$this.date_value.show",
                        expression=f"return ['{ConditionType.DATE_AFTER.value}', '{ConditionType.DATE_BEFORE.value}'].includes($this.condition_type.value)",
                    )
                ],
            ),
            atomicMg.param(
                "date_range",
                dynamics=[
                    DynamicsItem(
                        key="$this.date_range.show",
                        expression=f"return $this.condition_type.value == '{ConditionType.DATE_BETWEEN.value}'",
                    )
                ],
            ),
            atomicMg.param(
                "is_case_sensitive",
                dynamics=[
                    DynamicsItem(
                        key="$this.is_case_sensitive.show",
                        expression=f"return ['{ConditionType.EQUALS.value}', '{ConditionType.NOT_EQUALS.value}', '{ConditionType.CONTAINS.value}', '{ConditionType.NOT_CONTAINS.value}', '{ConditionType.STARTS_WITH.value}', '{ConditionType.ENDS_WITH.value}'].includes($this.condition_type.value)",
                    )
                ],
            ),
        ],
        outputList=[
            atomicMg.param(
                "data_filtered",
                types="List",
            ),
        ],
    )
    def filter_data_table(
        filter_type: FilterType = FilterType.COLUMN,
        row: int = 1,
        col: str = "A",
        condition_type: ConditionType = ConditionType.EQUALS,
        condition_value: str = "",
        date_value: str = "",
        date_range: str = "",
        is_case_sensitive: bool = True,
        is_save_filtered: bool = False,
    ) -> list:
        """
        过滤数据表格内容
        """
        if condition_type == ConditionType.DATE_BETWEEN:
            if not date_range:
                raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("日期范围不能为空"), "日期范围不能为空")
            elif len(date_range.split(",")) != 2:
                raise DATAFRAME_EXPECTION(
                    PARAMS_ERROR.format("日期范围格式错误，正确格式如：2023-01-01,2023-12-31"),
                    "日期范围格式错误，正确格式如：2023-01-01,2023-12-31",
                )
        col_index = col_to_index(col)
        data = []
        if filter_type == FilterType.COLUMN:
            if not col:
                raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("列号不能为空"), "列号不能为空")
            data = PyxlWrapper.read_column(col_index=col_index)
        elif filter_type == FilterType.ROW:
            row = validate_row_param(row)
            data = PyxlWrapper.read_row(row_index=row)
        else:
            data = PyxlWrapper.read_effective_area()

        data_filtered = filter_data(
            data=data,
            filter_type=filter_type,
            condition_type=condition_type,
            condition_value=condition_value,
            date_value=date_value,
            date_range=date_range,
            is_case_sensitive=is_case_sensitive,
        )

        if is_save_filtered:
            if filter_type == FilterType.COLUMN:
                PyxlWrapper.empty_column(col_index=col_index)
                DataTable.write_data(
                    write_type=WriteType.COLUMN,
                    col=col,
                    start_row=1,
                    data=data_filtered,
                    write_mode=WriteMode.OVERWRITE,
                )
            elif filter_type == FilterType.ROW:
                PyxlWrapper.empty_row(row_index=row)
                DataTable.write_data(
                    write_type=WriteType.ROW,
                    row=row,
                    data=data_filtered,
                    write_mode=WriteMode.OVERWRITE,
                )
            else:
                PyxlWrapper.clear_range(
                    range_str=f"A1:{index_to_col(PyxlWrapper.get_max_column() - 1)}{PyxlWrapper.get_max_row()}",
                )
                DataTable.write_data(
                    write_type=WriteType.AREA,
                    start_row=1,
                    start_col="A",
                    data=data_filtered,
                    write_mode=WriteMode.OVERWRITE,
                )

        return data_filtered

    @staticmethod
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param(
                "import_file_path",
                types="File",
                formType=AtomicFormTypeMeta(
                    type=AtomicFormType.INPUT_VARIABLE_PYTHON_FILE.value,
                    params={
                        "file_type": "file",
                        "filters": [".xlsx", ".xls", ".csv"],
                    },
                ),
            ),
            atomicMg.param(
                "sheet_name",
                types="Str",
                required=False,
            ),
            atomicMg.param(
                "password",
                types="Str",
                required=False,
            ),
            atomicMg.param("file_encoding"),
            atomicMg.param("first_row_is_header"),
            atomicMg.param(
                "csv_delimiter",
                required=False,
            ),
        ],
        outputList=[],
    )
    def import_data_table_from_file(
        import_file_path: str,
        sheet_name: str = None,  # type: ignore
        password: str = "",
        file_encoding: FileEncodingType = FileEncodingType.AUTO,
        first_row_is_header: bool = False,
        csv_delimiter: str = ",",
    ):
        """
        从指定文件导入数据表格
        :param import_file_path: 导入文件路径
        :param sheet_name: 工作表名称
        :param password: 打开密码(用于加密的Excel文件)
        :param file_encoding: 文件编码(仅对CSV生效)
        :param first_row_is_header: 首行是否作为列头(勾选后首行写入列头, 数据从第二行开始)
        :param csv_delimiter: CSV分隔符(仅对CSV生效, 默认逗号, 支持\t表示制表符)
        """
        if csv_delimiter == "\\t":
            csv_delimiter = "\t"
        if len(csv_delimiter) != 1:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format("CSV分隔符必须是单个字符"),
                "CSV分隔符必须是单个字符",
            )
        if not import_file_path:
            raise DATAFRAME_EXPECTION(IMPORT_FILE_ERROR_FORMAT.format("导入文件路径不能为空"), "导入文件路径不能为空")
        file_ext = os.path.splitext(import_file_path)[1].lower()
        if file_ext not in [".xlsx", ".xls", ".csv"]:
            raise DATAFRAME_EXPECTION(
                IMPORT_FILE_ERROR_FORMAT.format(""),
                "仅支持导入Excel(.xlsx, .xls)和CSV(.csv)文件",
            )
        if not os.path.exists(import_file_path):
            raise DATAFRAME_EXPECTION(
                IMPORT_FILE_ERROR_FORMAT.format(""),
                f"文件不存在: {import_file_path}",
            )

        if file_encoding == FileEncodingType.ANSI:
            encoding = "gbk"
        elif file_encoding == FileEncodingType.UTF8:
            encoding = "utf-8"
        elif file_encoding == FileEncodingType.UTF8_BOM:
            encoding = "utf-8-sig"
        else:
            encoding = None  # 自动识别

        temp_path = None
        import_path = import_file_path
        if password and file_ext in [".xlsx", ".xls"]:
            # 加密文件: 先解密到临时文件再导入
            temp_path = _decrypt_excel_to_temp_file(import_file_path, password)
            import_path = temp_path
        try:
            header_row = PyxlWrapper.fill_data_table_by_import_file(
                import_file_path=import_path,
                sheet_name=sheet_name,
                encoding=encoding,
                first_row_is_header=first_row_is_header,
                delimiter=csv_delimiter,
            )
            if header_row is not None:
                head_max_col = PyxlHeadWrapper.get_max_column()
                if head_max_col >= 1:
                    PyxlHeadWrapper.delete_cols(idx=1, amount=head_max_col)
                for col_idx, title in enumerate(header_row, start=1):
                    PyxlHeadWrapper.write_cell(row=1, col=col_idx, value=str(title) if title is not None else "")
                sync_data_table_head()
        finally:
            if temp_path and os.path.exists(temp_path):
                os.remove(temp_path)

    @staticmethod
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param(
                "export_dest_path",
                formType=AtomicFormTypeMeta(
                    type=AtomicFormType.INPUT_VARIABLE_PYTHON_FILE.value,
                    params={"file_type": "folder"},
                ),
            ),
            atomicMg.param(
                "export_file_name",
                required=False,
            ),
            atomicMg.param("csv_write_type"),
            atomicMg.param("file_encoding"),
            atomicMg.param(
                "csv_delimiter",
                required=False,
            ),
        ],
        outputList=[
            atomicMg.param(
                "export_file_path",
                types="Str",
            ),
        ],
    )
    def export_data_table_to_file(
        export_dest_path: str,
        export_file_name: str = "data_table",
        export_file_type: ExportFileType = ExportFileType.XLSX,
        is_overwrite: bool = True,
        csv_write_type: CsvWriteType = CsvWriteType.OVERWRITE,
        file_encoding: FileEncodingType = FileEncodingType.UTF8,
        csv_delimiter: str = ",",
    ) -> str:
        """
        导出数据表格到指定文件
        :param csv_write_type: CSV写入方式(追加/覆盖)
        :param file_encoding: 文件编码(仅对CSV/JSON生效)
        :param csv_delimiter: CSV分隔符(仅对CSV生效, 默认逗号, 支持\t表示制表符)
        """
        if csv_delimiter == "\\t":
            csv_delimiter = "\t"
        if len(csv_delimiter) != 1:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format("CSV分隔符必须是单个字符"),
                "CSV分隔符必须是单个字符",
            )
        if not export_dest_path:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("导出文件夹路径不能为空"), "导出文件夹路径不能为空")
        if not export_file_name:
            export_file_name = "data_table"
        if not os.path.exists(export_dest_path):
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format(f"导出文件夹路径不存在: {export_dest_path}"),
                "导出文件夹路径不存在",
            )
        if not is_overwrite:
            export_file_name = export_file_name + "_" + datetime.now().strftime("%Y%m%d%H%M%S")

        if file_encoding == FileEncodingType.ANSI:
            encoding = "gbk"
        elif file_encoding == FileEncodingType.UTF8_BOM:
            encoding = "utf-8-sig"
        else:
            encoding = "utf-8"

        file_path = os.path.join(export_dest_path, f"{export_file_name}.{export_file_type.value}")
        if export_file_type == ExportFileType.CSV:
            data = DataTable.read_data(
                read_type=ReadType.AREA,
                start_row=1,
                start_col="A",
                end_row=PyxlWrapper.get_max_row(),
                end_col=index_to_col(PyxlWrapper.get_max_column() - 1),
            )
            write_mode = "a" if (csv_write_type == CsvWriteType.APPEND and os.path.exists(file_path)) else "w"
            with open(file_path, write_mode, newline="", encoding=encoding) as csvfile:
                csv.writer(csvfile, delimiter=csv_delimiter).writerows(data)
        elif export_file_type == ExportFileType.JSON:
            data = DataTable.read_data(
                read_type=ReadType.AREA,
                start_row=1,
                start_col="A",
                end_row=PyxlWrapper.get_max_row(),
                end_col=index_to_col(PyxlWrapper.get_max_column() - 1),
            )
            with open(file_path, "w", encoding=encoding) as jsonfile:
                json.dump(data, jsonfile, indent=4)
        else:
            PyxlWrapper.export_to_file(file_path=file_path)
        return file_path

    @staticmethod
    @validate_cell
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param("start_row", required=False),
            atomicMg.param("start_col", required=False),
            atomicMg.param("end_col", required=False),
        ],
        outputList=[
            atomicMg.param("first_available_row", types="Int"),
        ],
    )
    def get_first_available_row(start_row: int = 1, start_col: str = "A", end_col: str = "-1") -> int:
        """
        获取列范围内第一个空白行的行号
        :param start_row: 起始行号
        :param start_col: 起始列号
        :param end_col: 结束列号, 不填则到已用最大列
        """
        if not start_col:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("开始列号不能为空"), "开始列号不能为空")
        if not start_row or int(start_row) < 1:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format("开始行号必须是大于0的正整数"),
                "开始行号必须是大于0的正整数",
            )
        start_col_index = col_to_index(start_col)
        if end_col is None or end_col in {"", "0", 0, -1, "-1"}:
            end_col_index = PyxlWrapper.get_max_column()
        else:
            end_col_index = col_to_index(end_col)
        if end_col_index < start_col_index:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format("结束列号不能小于开始列号"),
                "结束列号不能小于开始列号",
            )
        max_row = PyxlWrapper.get_max_row()
        for r in range(start_row, max_row + 1):
            is_blank = True
            for c in range(start_col_index, end_col_index + 1):
                value = PyxlWrapper.read_cell(row=r, col=c)
                if value is not None and value != "":
                    is_blank = False
                    break
            if is_blank:
                return r
        return max_row + 1

    @staticmethod
    @validate_cell
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param("start_col", required=False),
            atomicMg.param("start_row", required=False),
            atomicMg.param("end_row", required=False),
        ],
        outputList=[
            atomicMg.param("first_available_col", types="Str"),
        ],
    )
    def get_first_available_col(start_col: str = "A", start_row: int = 1, end_row: int = -1) -> str:
        """
        获取行范围内第一个空白列的列标
        :param start_col: 起始列号
        :param start_row: 起始行号
        :param end_row: 结束行号, -1表示最后一行(默认), 0或不填为已用最大行
        """
        if not start_col:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("开始列号不能为空"), "开始列号不能为空")
        if not start_row or int(start_row) < 1:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format("开始行号必须是大于0的正整数"),
                "开始行号必须是大于0的正整数",
            )
        end_row = normalize_end_row(end_row)
        if int(end_row) < int(start_row):
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format("结束行号不能小于开始行号"),
                "结束行号不能小于开始行号",
            )
        start_col_index = col_to_index(start_col)
        max_col = PyxlWrapper.get_max_column()
        for c in range(start_col_index, max_col + 1):
            is_blank = True
            for r in range(int(start_row), int(end_row) + 1):
                value = PyxlWrapper.read_cell(row=r, col=c)
                if value is not None and value != "":
                    is_blank = False
                    break
            if is_blank:
                return index_to_col(c - 1)
        return index_to_col(max_col)

    @staticmethod
    @validate_cell
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param("col", required=False),
            atomicMg.param("start_row", required=False),
        ],
        outputList=[
            atomicMg.param("available_row", types="Int"),
        ],
    )
    def get_first_available_row_by_col(col: str = "A", start_row: int = 1) -> int:
        """
        获取指定列第一个空白单元格的行号
        :param col: 列号
        :param start_row: 起始行号
        """
        if not col:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("列号不能为空"), "列号不能为空")
        if not start_row or int(start_row) < 1:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format("开始行号必须是大于0的正整数"),
                "开始行号必须是大于0的正整数",
            )
        col_index = col_to_index(col)
        max_row = PyxlWrapper.get_max_row()
        for r in range(int(start_row), max_row + 1):
            value = PyxlWrapper.read_cell(row=r, col=col_index)
            if value is None or value == "":
                return r
        return max_row + 1

    @staticmethod
    @validate_cell
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param("row", required=False),
            atomicMg.param("start_col", required=False),
        ],
        outputList=[
            atomicMg.param("available_col", types="Str"),
        ],
    )
    def get_first_available_col_by_row(row: int = 1, start_col: str = "A") -> str:
        """
        获取指定行第一个空白单元格的列标
        :param row: 行号
        :param start_col: 起始列号
        """
        if not row or int(row) < 1:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format("行号必须是大于0的正整数"),
                "行号必须是大于0的正整数",
            )
        if not start_col:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("开始列号不能为空"), "开始列号不能为空")
        start_col_index = col_to_index(start_col)
        max_col = PyxlWrapper.get_max_column()
        for c in range(start_col_index, max_col + 1):
            value = PyxlWrapper.read_cell(row=int(row), col=c)
            if value is None or value == "":
                return index_to_col(c - 1)
        return index_to_col(max_col)

    @staticmethod
    @validate_cell
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param("start_row", required=False),
            atomicMg.param("cols", required=False),
            atomicMg.param("keep_first", required=False),
        ],
        outputList=[
            atomicMg.param("removed_count", types="Int"),
        ],
    )
    def remove_duplicate_rows(start_row: int = 1, cols: str = "", keep_first: bool = True) -> int:
        """
        删除数据表格中的重复行
        :param start_row: 起始行号
        :param cols: 比较列, 多列用逗号分隔(如A,C), 不填则比较整行
        :param keep_first: 是否保留第一次出现的重复行
        """
        if not start_row or int(start_row) < 1:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format("开始行号必须是大于0的正整数"),
                "开始行号必须是大于0的正整数",
            )
        if cols is None or cols == "":
            col_indexes = list(range(1, PyxlWrapper.get_max_column() + 1))
        else:
            col_indexes = parse_col_numbers(cols)
        # max_row可能含delete_rows残留的幻影空行, 以上最后一个非空行为界
        max_row = last_nonempty_row()

        def row_key(r):
            return tuple(
                "" if value is None else str(value)
                for value in (PyxlWrapper.read_cell(row=r, col=c) for c in col_indexes)
            )

        rows_to_delete = []
        seen = set()
        if keep_first:
            for r in range(int(start_row), max_row + 1):
                key = row_key(r)
                if key in seen:
                    rows_to_delete.append(r)
                else:
                    seen.add(key)
        else:
            for r in range(max_row, int(start_row) - 1, -1):
                key = row_key(r)
                if key in seen:
                    rows_to_delete.append(r)
                else:
                    seen.add(key)
        # 从大到小删除避免行号位移
        for r in sorted(rows_to_delete, reverse=True):
            PyxlWrapper.delete_rows(idx=r, amount=1)
        return len(rows_to_delete)

    @staticmethod
    @validate_cell
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param("sort_cols", required=False),
            atomicMg.param("sort_orders", required=False),
            atomicMg.param("start_row", required=False),
            atomicMg.param("end_row", required=False),
            atomicMg.param("has_header", required=False),
        ],
        outputList=[],
    )
    def sort_data_table(
        sort_cols: str = "A",
        sort_orders: str = "ascending",
        start_row: int = 1,
        end_row: int = -1,
        has_header: bool = False,
    ):
        """
        排序数据表格
        :param sort_cols: 排序列, 多列用逗号分隔(如A,C)
        :param sort_orders: 排序方式, 逗号分隔与排序列一一对应(ascending/descending)
        :param start_row: 起始行号
        :param end_row: 结束行号, -1表示最后一行(默认), 0或不填为已用最大行
        :param has_header: 首行是否为表头(不参与排序)
        """
        if not sort_cols:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("排序列不能为空"), "排序列不能为空")
        if not start_row or int(start_row) < 1:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format("开始行号必须是大于0的正整数"),
                "开始行号必须是大于0的正整数",
            )
        # 保序解析排序列(主键在前), 不使用parse_col_numbers(其会去重升序, 破坏主次键优先级)
        col_indexes = []
        for part in str(sort_cols).split(","):
            part = part.strip()
            if not part:
                continue
            idx = col_to_index(part)
            if idx not in col_indexes:
                col_indexes.append(idx)
        order_list = [order.strip() for order in str(sort_orders).split(",") if order.strip()]
        if len(order_list) != len(col_indexes):
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format("排序方式数量必须与排序列数量一致"),
                "排序方式数量必须与排序列数量一致",
            )
        for order in order_list:
            if order not in (SortOrder.ASCENDING.value, SortOrder.DESCENDING.value):
                raise DATAFRAME_EXPECTION(
                    PARAMS_ERROR.format(f"无效的排序方式: {order}"),
                    f"无效的排序方式: {order}，仅支持ascending/descending",
                )
        end_row = normalize_end_row(end_row)
        # max_row可能含delete_rows残留的幻影空行, 截断到最后一个非空行, 避免空行被排到数据前面
        end_row = min(int(end_row), last_nonempty_row())
        data_start_row = int(start_row) + 1 if has_header else int(start_row)
        if data_start_row > int(end_row):
            return
        # 整行重排: 读取全部已用列, 保证行数据完整性
        max_col = PyxlWrapper.get_max_column()
        rows = [
            [PyxlWrapper.read_cell(row=r, col=c) for c in range(1, max_col + 1)]
            for r in range(data_start_row, int(end_row) + 1)
        ]

        def sort_key_value(value):
            # 规范化排序键: None最前, 数值次之, 字符串最后
            if value is None:
                return (0, "")
            if isinstance(value, (int, float)):
                return (1, value)
            return (2, str(value))

        # 多关键字排序: 利用稳定排序从最后一列往前依次排序
        for col_index, order in reversed(list(zip(col_indexes, order_list))):
            offset = col_index - 1
            reverse = order == SortOrder.DESCENDING.value
            rows.sort(key=lambda row_values: sort_key_value(row_values[offset]), reverse=reverse)

        for i, row_values in enumerate(rows):
            for offset, value in enumerate(row_values):
                # 直接赋值: cell(value=None)是空操作, 无法清空排序后空出的单元格
                PyxlWrapper.sheet.cell(row=data_start_row + i, column=offset + 1).value = value
        # 直接写 sheet 绕过了 wrapper 的边界缓存, 手动失效
        PyxlWrapper.invalidate_bounds()

    @staticmethod
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param(
                "validate_range",
                required=True,
            ),
            atomicMg.param(
                "validate_type",
                formType=AtomicFormTypeMeta(AtomicFormType.SELECT.value),
            ),
            atomicMg.param(
                "validate_operator",
                formType=AtomicFormTypeMeta(AtomicFormType.SELECT.value),
            ),
            atomicMg.param("formula1", required=False),
            atomicMg.param("formula2", required=False),
            atomicMg.param("allow_blank", required=False),
            atomicMg.param("error_msg", required=False),
        ],
        outputList=[],
    )
    def add_data_validation(
        validate_range: str = "A1:B10",
        validate_type: ValidateType = ValidateType.WHOLE,
        validate_operator: ValidateOperator = ValidateOperator.BETWEEN,
        formula1: str = "",
        formula2: str = "",
        allow_blank: bool = True,
        error_msg: str = "",
    ):
        """
        添加数据验证
        :param validate_range: 验证区域(如A1:B10或A1)
        :param validate_type: 验证类型
        :param validate_operator: 验证操作符(list/custom类型不需要)
        :param formula1: 条件1(如最小值或列表值)
        :param formula2: 条件2(如最大值)
        :param allow_blank: 是否允许空值
        :param error_msg: 错误提示信息
        """
        if not validate_range:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("验证区域不能为空"), "验证区域不能为空")
        validate_range = str(validate_range).strip()
        if not re.match(r"^[A-Za-z]{1,3}[1-9]\d*(:[A-Za-z]{1,3}[1-9]\d*)?$", validate_range):
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format(f"验证区域格式错误: {validate_range}"),
                f"验证区域格式错误: {validate_range}，正确格式如 A1:B10",
            )
        validate_type = _to_enum(validate_type, ValidateType)
        validate_operator = _to_enum(validate_operator, ValidateOperator)
        dv_kwargs = {
            "type": validate_type.value,
            "allowBlank": allow_blank,
            "showErrorMessage": True,
            "errorTitle": "输入无效",
            "error": error_msg,
        }
        if validate_type not in (ValidateType.LIST, ValidateType.CUSTOM):
            dv_kwargs["operator"] = validate_operator.value
        dv = DataValidation(**dv_kwargs)
        if formula1:
            dv.formula1 = str(formula1)
        if formula2:
            dv.formula2 = str(formula2)
        dv.add(validate_range)
        PyxlWrapper.sheet.add_data_validation(dv)

    @staticmethod
    @validate_cell
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param(
                "format_type",
                formType=AtomicFormTypeMeta(AtomicFormType.SELECT.value),
            ),
            atomicMg.param("row", required=False),
            atomicMg.param("col", required=False),
            atomicMg.param("start_row", required=False),
            atomicMg.param("start_col", required=False),
            atomicMg.param("end_row", required=False),
            atomicMg.param("end_col", required=False),
            atomicMg.param("font_name", required=False),
            atomicMg.param("font_size", required=False),
            atomicMg.param("bold", required=False),
            atomicMg.param("italic", required=False),
            atomicMg.param(
                "underline",
                formType=AtomicFormTypeMeta(AtomicFormType.SELECT.value),
                required=False,
            ),
            atomicMg.param("font_color", required=False),
            atomicMg.param("bg_color", required=False),
            atomicMg.param(
                "h_align",
                formType=AtomicFormTypeMeta(AtomicFormType.SELECT.value),
                required=False,
            ),
            atomicMg.param(
                "v_align",
                formType=AtomicFormTypeMeta(AtomicFormType.SELECT.value),
                required=False,
            ),
            atomicMg.param("wrap_text", required=False),
            atomicMg.param(
                "border_style",
                formType=AtomicFormTypeMeta(AtomicFormType.SELECT.value),
                required=False,
            ),
        ],
        outputList=[],
    )
    def set_format(
        format_type: BaseOperateType = BaseOperateType.AREA,
        row: int = 1,
        col: str = "A",
        start_row: int = 1,
        start_col: str = "A",
        end_row: int = -1,
        end_col: str = "-1",
        font_name: str = "",
        font_size: float = 0,
        bold: bool = False,
        italic: bool = False,
        underline: UnderlineType = UnderlineType.NONE,
        font_color: str = "",
        bg_color: str = "",
        h_align: HAlignType = HAlignType.NONE,
        v_align: VAlignType = VAlignType.NONE,
        wrap_text: bool = False,
        border_style: BorderStyleType = BorderStyleType.NONE,
    ):
        """
        设置数据表格格式(字体/填充/对齐/边框)
        :param format_type: 格式设置方式(单元格/行/列/区域)
        :param font_name: 字体名称, 不填则不设置
        :param font_size: 字号, 0或不填则不设置
        :param bold: 是否加粗
        :param italic: 是否斜体
        :param underline: 下划线类型
        :param font_color: 字体颜色(如FF0000), 不填则不设置
        :param bg_color: 背景颜色(如FFFF00), 不填则不设置
        :param h_align: 水平对齐方式
        :param v_align: 垂直对齐方式
        :param wrap_text: 是否自动换行
        :param border_style: 边框样式
        """
        min_row, max_row, min_col, max_col = _resolve_format_area(
            format_type, row, col, start_row, start_col, end_row, end_col
        )
        underline = _to_enum(underline, UnderlineType)
        border_style = _to_enum(border_style, BorderStyleType)
        font_kwargs = {}
        if font_name:
            font_kwargs["name"] = font_name
        if font_size and font_size > 0:
            font_kwargs["size"] = font_size
        if bold:
            font_kwargs["bold"] = True
        if italic:
            font_kwargs["italic"] = True
        if underline != UnderlineType.NONE:
            font_kwargs["underline"] = underline.value
        if font_color:
            font_kwargs["color"] = font_color
        align_kwargs = {}
        h_align = _to_enum(h_align, HAlignType)
        v_align = _to_enum(v_align, VAlignType)
        if h_align != HAlignType.NONE:
            align_kwargs["horizontal"] = h_align.value
        if v_align != VAlignType.NONE:
            align_kwargs["vertical"] = v_align.value
        if wrap_text:
            align_kwargs["wrap_text"] = True
        side = None
        if border_style != BorderStyleType.NONE:
            side = Side(style=border_style.value)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = PyxlWrapper.sheet.cell(row=r, column=c)
                if font_kwargs:
                    # 合并现有字体属性, 未指定的项保持原值(不填则不设置)
                    cell.font = Font(
                        name=font_kwargs.get("name", cell.font.name),
                        size=font_kwargs.get("size", cell.font.size),
                        bold=font_kwargs.get("bold", cell.font.bold),
                        italic=font_kwargs.get("italic", cell.font.italic),
                        underline=font_kwargs.get("underline", cell.font.underline),
                        color=font_kwargs.get("color", cell.font.color),
                    )
                if bg_color:
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                if align_kwargs:
                    # 合并现有对齐属性, 未指定的项保持原值
                    cell.alignment = Alignment(
                        horizontal=align_kwargs.get("horizontal", cell.alignment.horizontal),
                        vertical=align_kwargs.get("vertical", cell.alignment.vertical),
                        wrap_text=align_kwargs.get("wrap_text", cell.alignment.wrap_text),
                    )
                if side is not None:
                    cell.border = Border(left=side, right=side, top=side, bottom=side)

    @staticmethod
    @validate_cell
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param(
                "format_type",
                formType=AtomicFormTypeMeta(AtomicFormType.SELECT.value),
            ),
            atomicMg.param("row", required=False),
            atomicMg.param("col", required=False),
            atomicMg.param("start_row", required=False),
            atomicMg.param("start_col", required=False),
            atomicMg.param("end_row", required=False),
            atomicMg.param("end_col", required=False),
        ],
        outputList=[],
    )
    def clear_format(
        format_type: BaseOperateType = BaseOperateType.AREA,
        row: int = 1,
        col: str = "A",
        start_row: int = 1,
        start_col: str = "A",
        end_row: int = -1,
        end_col: str = "-1",
    ):
        """
        清除数据表格格式(恢复常规样式)
        :param format_type: 格式清除方式(单元格/行/列/区域)
        """
        min_row, max_row, min_col, max_col = _resolve_format_area(
            format_type, row, col, start_row, start_col, end_row, end_col
        )
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                PyxlWrapper.sheet.cell(row=r, column=c).style = "Normal"

    @staticmethod
    @validate_cell
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param("row", required=False),
            atomicMg.param("height", required=False),
        ],
        outputList=[],
    )
    def set_row_height(row: str = "1", height: float = 20.0):
        """
        设置行高
        :param row: 行号, 支持批量(如1,3,5:7), 支持负数(-1表示最后一行)
        :param height: 行高
        """
        if not row:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("行号不能为空"), "行号不能为空")
        if height is None or float(height) < 0:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("行高必须大于等于0"), "行高必须大于等于0")
        for r in parse_row_numbers(row):
            PyxlWrapper.sheet.row_dimensions[r].height = float(height)

    @staticmethod
    @validate_cell
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param("col", required=False),
            atomicMg.param("width", required=False),
        ],
        outputList=[],
    )
    def set_column_width(col: str = "A", width: float = 15.0):
        """
        设置列宽
        :param col: 列号, 支持批量(如A,C,E:G), 支持负数(-1表示最后一列)
        :param width: 列宽
        """
        if not col:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("列号不能为空"), "列号不能为空")
        if width is None or float(width) < 0:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("列宽必须大于等于0"), "列宽必须大于等于0")
        for c in parse_col_numbers(col):
            PyxlWrapper.sheet.column_dimensions[index_to_col(c - 1)].width = float(width)

    @staticmethod
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param(
                "target_type",
                formType=AtomicFormTypeMeta(AtomicFormType.SELECT.value),
            ),
            atomicMg.param("target", required=False),
            atomicMg.param("is_hidden", required=False),
        ],
        outputList=[],
    )
    def set_row_col_hidden(target_type: HideTargetType = HideTargetType.ROW, target: str = "1", is_hidden: bool = True):
        """
        隐藏或显示行/列
        :param target_type: 目标类型(行/列)
        :param target: 行号或列号, 支持批量(如1,3,5:7或A,C,E:G), 支持负数(-1表示最后一行/列)
        :param is_hidden: 是否隐藏
        """
        target_type = _to_enum(target_type, HideTargetType)
        if not target:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format(f"目标{target_type.value}不能为空"),
                f"目标{target_type.value}不能为空",
            )
        if target_type == HideTargetType.ROW:
            for r in parse_row_numbers(target):
                PyxlWrapper.sheet.row_dimensions[r].hidden = is_hidden
        else:
            for c in parse_col_numbers(target):
                PyxlWrapper.sheet.column_dimensions[index_to_col(c - 1)].hidden = is_hidden

    @staticmethod
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param("sheet_name", required=True),
            atomicMg.param("position", required=False),
        ],
        outputList=[],
    )
    def add_sheet(sheet_name: str = "Sheet2", position: int = 0):
        """
        新增工作表
        :param sheet_name: 新工作表名称
        :param position: 插入位置, 0表示添加到最后
        """
        if not sheet_name:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("工作表名称不能为空"), "工作表名称不能为空")
        if sheet_name in PyxlWrapper.workbook.sheetnames:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format(f"工作表[{sheet_name}]已存在"),
                f"工作表[{sheet_name}]已存在",
            )
        index = int(position) if position and int(position) > 0 else None
        PyxlWrapper.add_sheet(title=sheet_name, index=index)
        PyxlWrapper.switch_sheet(sheet_name)

    @staticmethod
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param("sheet_name", required=True),
        ],
        outputList=[
            atomicMg.param("current_sheet", types="Str"),
        ],
    )
    def activate_sheet(sheet_name: str = "") -> str:
        """
        激活工作表(切换当前操作的工作表)
        :param sheet_name: 工作表名称
        """
        if not sheet_name:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("工作表名称不能为空"), "工作表名称不能为空")
        if sheet_name not in PyxlWrapper.workbook.sheetnames:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format(f"工作表[{sheet_name}]不存在"),
                f"工作表[{sheet_name}]不存在",
            )
        PyxlWrapper.switch_sheet(sheet_name)
        return PyxlWrapper.sheet.title

    @staticmethod
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param("source_sheet_name", required=True),
            atomicMg.param("new_sheet_name", required=True),
        ],
        outputList=[],
    )
    def copy_sheet(source_sheet_name: str = "", new_sheet_name: str = ""):
        """
        复制工作表
        :param source_sheet_name: 源工作表名称
        :param new_sheet_name: 新工作表名称
        """
        if not source_sheet_name:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("源工作表名称不能为空"), "源工作表名称不能为空")
        if not new_sheet_name:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("新工作表名称不能为空"), "新工作表名称不能为空")
        sheet_names = PyxlWrapper.workbook.sheetnames
        if source_sheet_name not in sheet_names:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format(f"源工作表[{source_sheet_name}]不存在"),
                f"源工作表[{source_sheet_name}]不存在",
            )
        if new_sheet_name in sheet_names:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format(f"工作表[{new_sheet_name}]已存在"),
                f"工作表[{new_sheet_name}]已存在",
            )
        PyxlWrapper.copy_sheet(source_sheet_name=source_sheet_name, new_sheet_name=new_sheet_name)
        PyxlWrapper.switch_sheet(new_sheet_name)

    @staticmethod
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param("sheet_name", required=True),
        ],
        outputList=[],
    )
    def delete_sheet(sheet_name: str = ""):
        """
        删除工作表(至少保留一个工作表)
        :param sheet_name: 工作表名称
        """
        if not sheet_name:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("工作表名称不能为空"), "工作表名称不能为空")
        sheet_names = PyxlWrapper.workbook.sheetnames
        if sheet_name not in sheet_names:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format(f"工作表[{sheet_name}]不存在"),
                f"工作表[{sheet_name}]不存在",
            )
        if len(sheet_names) <= 1:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format("数据表格至少需要保留一个工作表"),
                "数据表格至少需要保留一个工作表",
            )
        is_current = PyxlWrapper.sheet.title == sheet_name
        PyxlWrapper.delete_sheet(sheet_name)
        if is_current:
            # 删除的是当前表, 切换到工作簿的活动工作表
            PyxlWrapper.sheet = PyxlWrapper.workbook.active
            PyxlWrapper.invalidate_bounds()

    @staticmethod
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param("old_sheet_name", required=True),
            atomicMg.param("new_sheet_name", required=True),
        ],
        outputList=[],
    )
    def rename_sheet(old_sheet_name: str = "", new_sheet_name: str = ""):
        """
        重命名工作表
        :param old_sheet_name: 原工作表名称
        :param new_sheet_name: 新工作表名称
        """
        if not old_sheet_name:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("原工作表名称不能为空"), "原工作表名称不能为空")
        if not new_sheet_name:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("新工作表名称不能为空"), "新工作表名称不能为空")
        sheet_names = PyxlWrapper.workbook.sheetnames
        if old_sheet_name not in sheet_names:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format(f"工作表[{old_sheet_name}]不存在"),
                f"工作表[{old_sheet_name}]不存在",
            )
        if new_sheet_name in sheet_names:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format(f"工作表[{new_sheet_name}]已存在"),
                f"工作表[{new_sheet_name}]已存在",
            )
        PyxlWrapper.rename_sheet(old_name=old_sheet_name, new_name=new_sheet_name)

    @staticmethod
    @atomicMg.atomic(
        "DataTable",
        inputList=[],
        outputList=[
            atomicMg.param("sheet_names", types="List"),
            atomicMg.param("current_sheet", types="Str"),
        ],
    )
    def get_sheet_names() -> list:
        """
        获取全部工作表名称与当前工作表名称
        """
        return PyxlWrapper.get_sheet_names(), PyxlWrapper.sheet.title

    @staticmethod
    @auto_save
    @atomicMg.atomic(
        "DataTable",
        inputList=[
            atomicMg.param(
                "export_dest_path",
                formType=AtomicFormTypeMeta(
                    type=AtomicFormType.INPUT_VARIABLE_PYTHON_FILE.value,
                    params={"file_type": "folder"},
                ),
            ),
            atomicMg.param(
                "export_file_name",
                required=False,
            ),
            atomicMg.param(
                "is_overwrite",
                required=False,
            ),
        ],
        outputList=[
            atomicMg.param(
                "pdf_path",
                types="Str",
            ),
        ],
    )
    def export_data_table_to_pdf(
        export_dest_path: str,
        export_file_name: str = "data_table",
        is_overwrite: bool = True,
    ) -> str:
        """
        导出数据表格为PDF文件(需要安装Excel/WPS或LibreOffice)
        :param export_dest_path: 导出目标文件夹路径
        :param export_file_name: 导出文件名(不含扩展名)
        :param is_overwrite: 是否覆盖同名文件, 否则自动重命名
        """
        if not export_dest_path:
            raise DATAFRAME_EXPECTION(PARAMS_ERROR.format("导出文件夹路径不能为空"), "导出文件夹路径不能为空")
        if not export_file_name:
            export_file_name = "data_table"
        if not os.path.exists(export_dest_path):
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format(f"导出文件夹路径不存在: {export_dest_path}"),
                "导出文件夹路径不存在",
            )
        if not is_overwrite:
            export_file_name = export_file_name + "_" + datetime.now().strftime("%Y%m%d%H%M%S")
        pdf_path = os.path.join(export_dest_path, f"{export_file_name}.pdf")
        if os.path.exists(pdf_path) and not is_overwrite:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format(f"文件已存在: {pdf_path}"),
                f"文件已存在: {pdf_path}",
            )

        # 先保存确保磁盘上的xlsx文件为最新
        PyxlWrapper.save(path=_xlsx_file_path)
        xlsx_path = os.path.abspath(_xlsx_file_path)
        pdf_abs_path = os.path.abspath(pdf_path)
        errors = []
        converted = False
        if sys.platform == "win32":
            # Windows: 按序尝试 Excel/WPS 的COM组件
            for prog_id in ("Excel.Application", "Ket.Application", "et.Application"):
                try:
                    import win32com.client
                except ImportError as e:
                    errors.append(f"win32com: {e}")
                    break
                try:
                    app = win32com.client.DispatchEx(prog_id)
                except Exception as e:
                    errors.append(f"{prog_id}: {e}")
                    continue
                try:
                    app.Visible = False
                    wb = app.Workbooks.Open(xlsx_path)
                    try:
                        wb.ExportAsFixedFormat(0, pdf_abs_path)
                        converted = True
                    finally:
                        wb.Close(False)
                except Exception as e:
                    errors.append(f"{prog_id}: {e}")
                finally:
                    try:
                        app.Quit()
                    except Exception:
                        pass
                if converted:
                    break
        if not converted:
            # 非Windows或COM失败: 使用LibreOffice转换
            soffice = shutil.which("soffice")
            if soffice:
                try:
                    subprocess.run(
                        [soffice, "--headless", "--convert-to", "pdf", "--outdir", export_dest_path, xlsx_path],
                        check=True,
                        timeout=120,
                    )
                    generated = os.path.join(
                        export_dest_path,
                        os.path.splitext(os.path.basename(xlsx_path))[0] + ".pdf",
                    )
                    if os.path.abspath(generated) != pdf_abs_path and os.path.exists(generated):
                        os.replace(generated, pdf_abs_path)
                    if os.path.exists(pdf_abs_path):
                        converted = True
                    else:
                        errors.append("LibreOffice: 转换后未生成PDF文件")
                except Exception as e:
                    errors.append(f"LibreOffice: {e}")
        if not converted:
            raise DATAFRAME_EXPECTION(
                PARAMS_ERROR.format(f"导出PDF失败: {'; '.join(errors) if errors else '未找到可用的转换工具'}"),
                "导出PDF失败，需要安装Excel/WPS或LibreOffice",
            )
        return pdf_path
