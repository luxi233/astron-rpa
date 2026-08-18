"""excel_service 单元格更新合并防抖冒烟测试。

覆盖:
1. 快速连续 update_cells: 合并为一次落盘(计数 _apply_updates 调用)
2. 读接口自动 flush: pending 更新在读文件时立即可见
3. 定时器静默期触发: 200ms 无新更新后自动落盘
4. flush 异常容错: pending 指向已删除文件时不抛错、不影响其他文件

运行: cd engine/servers/astronverse-scheduler && uv run --with openpyxl python tests/smoke/smoke_excel_debounce.py
"""

import os
import sys
import tempfile
import time
from unittest.mock import patch

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
sys.path.insert(0, os.path.join(ROOT, "src"))

from astronverse.scheduler.core.datatable.excel_service import ExcelService  # noqa: E402

PASS = 0
FAIL = 0


def check(name: str, cond: bool, detail: str = ""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"[PASS] {name}")
    else:
        FAIL += 1
        print(f"[FAIL] {name} {detail}")


td = tempfile.mkdtemp()
svc = ExcelService(td)

# 建两个文件
f1 = svc.create_file("t1")
f2 = svc.create_file("t2")
check("准备: 文件已创建", os.path.exists(f1) and os.path.exists(f2))

apply_calls = []
orig_apply = (
    ExcelService._apply_updates.__func__
    if hasattr(ExcelService._apply_updates, "__func__")
    else ExcelService._apply_updates
)


def counting_apply(file_path, updates):
    apply_calls.append((file_path, len(updates)))
    return orig_apply(file_path, updates)


# ---------- 1. 快速连写合并 ----------
with patch.object(ExcelService, "_apply_updates", staticmethod(counting_apply)):
    for i in range(5):
        svc.update_cells("t1", [{"sheet": "Sheet", "row": 0, "col": i, "value": f"v{i}"}])
    # 防抖窗口内: 不落盘
    check("防抖: 200ms窗口内未落盘", len(apply_calls) == 0 and len(ExcelService._pending_updates) == 1)

    # 等静默期过定时器触发
    time.sleep(0.5)

check("防抖: 静默期后合并落盘1次", len(apply_calls) == 1 and apply_calls[0][1] == 5, str(apply_calls))
check("防抖: pending清空", not ExcelService._pending_updates and ExcelService._flush_timer is None)

# 文件内容验证: 5个单元格全部写入
import openpyxl  # noqa: E402

wb = openpyxl.load_workbook(f1)
ws = wb["Sheet"]
vals = [ws.cell(row=1, column=i + 1).value for i in range(5)]
wb.close()
check("落盘: 5格全部写入", vals == [f"v{i}" for i in range(5)], str(vals))

# ---------- 2. 读接口自动 flush ----------
apply_calls.clear()
svc.update_cells("t1", [{"sheet": "Sheet", "row": 1, "col": 0, "value": "read-flush"}])
data = svc.read_file("t1")  # 入口应先flush
check("读flush: pending已清空", not ExcelService._pending_updates)
cell = data["sheets"][0]["data"][1][0]
check("读flush: 读到pending中的值", cell == "read-flush", f"got {cell}")

# ---------- 3. 跨实例状态共享(get_excel_service 每请求新建) ----------
svc2 = ExcelService(td)
svc2.update_cells("t2", [{"sheet": "Sheet", "row": 0, "col": 0, "value": "cross-inst"}])
data2 = svc.read_file("t2")  # 另一实例读取, 仍应看到实例2的pending
check("跨实例: 类级pending共享", data2["sheets"][0]["data"][0][0] == "cross-inst")

# ---------- 4. flush 异常容错 ----------
# 人为塞入指向已删除文件的pending, flush 不应抛错
ExcelService._pending_updates[os.path.join(td, "gone.xlsx")] = [{"sheet": "S", "row": 0, "col": 0, "value": 1}]
svc.update_cells("t2", [{"sheet": "Sheet", "row": 1, "col": 0, "value": "survivor"}])
try:
    svc.flush_pending()
    check("容错: flush不抛异常", True)
except Exception as e:
    check("容错: flush不抛异常", False, str(e))
wb = openpyxl.load_workbook(f2)
ws2 = wb["Sheet"]
wb.close()
check("容错: 其他文件正常落盘", ws2.cell(row=2, column=1).value == "survivor")

print(f"\n结果: {PASS} 通过, {FAIL} 失败")
sys.exit(1 if FAIL else 0)
