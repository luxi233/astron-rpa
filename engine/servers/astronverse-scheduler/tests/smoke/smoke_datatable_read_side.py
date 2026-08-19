"""scheduler 数据表格读侧稳定性测试。

覆盖前端显示链路(scheduler 侧)的稳定性:
E. ExcelService: 读裁剪/防抖合并落盘/pending一致性/原子写/并发读写零失败/删除清理
F. AsyncFileWatcher: watchdog线程->事件循环的线程安全调度/防抖延迟/pause忽略

运行: cd engine/servers/astronverse-scheduler && .venv/bin/python tests/smoke/smoke_datatable_read_side.py
"""

import asyncio
import os
import shutil
import sys
import tempfile
import threading
import time

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "src"))

from astronverse.scheduler.core.datatable.excel_service import ExcelService
from astronverse.scheduler.core.datatable.file_watcher import AsyncFileWatcher

PASS = 0
FAIL = 0


def check(name: str, cond: bool, detail: str = ""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"[PASS] {name}")
    else:
        FAIL += 1
        print(f"[FAIL] {name} {detail}")


# realpath 消除 macOS /var -> /private/var symlink: watchdog(FSEvents)报的是真实路径,
# normpath 不解析 symlink, 路径不匹配会导致事件被 _should_process 过滤
tmpdir = os.path.realpath(tempfile.mkdtemp(prefix="dt_sched_"))
svc = ExcelService(resource_dir=tmpdir)

# ============================================================
# E. ExcelService
# ============================================================
print("\n========== E. ExcelService ==========")

# E1 创建 + 尾部空行/列裁剪
svc.create_file("t1")
svc.update_cells(
    "t1",
    [
        {"sheet": "Sheet", "row": 0, "col": 0, "value": "a1"},
        {"sheet": "Sheet", "row": 0, "col": 1, "value": "b1"},
        {"sheet": "Sheet", "row": 1, "col": 0, "value": "a2"},
    ],
)
svc.flush_pending()
# 直接写入尾部空值撑大 dimension, 验证 read_file 收敛
from openpyxl import load_workbook

fp = svc.get_file_path("t1")
w = load_workbook(fp)
w.active.cell(row=50, column=20, value=None)
w.active.cell(row=3, column=2, value=None)  # 注意: value=None 不落盘, 手动造空串
w.save(fp)
w.close()
data = svc.read_file("t1")
sheet = data["sheets"][0]
check(
    "E1 read_file 裁剪尾部空行/列",
    sheet["max_row"] == 2 and sheet["max_column"] == 2,
    f"got {sheet['max_row']}x{sheet['max_column']}",
)

# E2 防抖合并 + read 前自动 flush 一致性
for i in range(10):
    svc.update_cells("t1", [{"sheet": "Sheet", "row": 5 + i, "col": 0, "value": f"deb-{i}"}])
data = svc.read_file("t1")  # read 内部先 flush_pending
vals = [data["sheets"][0]["data"][5 + i][0] if len(data["sheets"][0]["data"]) > 5 + i else None for i in range(10)]
check("E2 防抖合并写读一致(10行全可见)", vals == [f"deb-{i}" for i in range(10)], f"got {vals}")

# E3 空值写入已用区域内可清空
svc.update_cells("t1", [{"sheet": "Sheet", "row": 5, "col": 0, "value": None}])
svc.flush_pending()
data = svc.read_file("t1")
v = data["sheets"][0]["data"][5][0] if len(data["sheets"][0]["data"]) > 5 else None
check("E3 区域内空值清空生效", v is None, f"got {v}")

# E4 空值写入已用区域外不撑大边界
svc.update_cells("t1", [{"sheet": "Sheet", "row": 999, "col": 0, "value": None}])
svc.flush_pending()
data = svc.read_file("t1")
check("E4 区域外空值不污染边界", data["sheets"][0]["max_row"] <= 15, f"got {data['sheets'][0]['max_row']}")

# E5 write_file 全量覆盖(原子)后可读
svc.write_file("t2", {"sheets": [{"name": "S1", "data": [["x", "y"], ["1", "2"]]}], "active_sheet": "S1"})
data = svc.read_file("t2")
check("E5 write_file 后内容正确", data["sheets"][0]["data"][:2] == [["x", "y"], ["1", "2"]])

# E6 delete_file 后 read 抛 FileNotFoundError(不复活)
svc.delete_file("t2")
try:
    svc.read_file("t2")
    check("E6 删除后读报错不复活", False)
except FileNotFoundError:
    check("E6 删除后读报错不复活", True)
check("E6b 文件已删除", not svc.file_exists("t2"))

# E7 并发 update_cells(多线程) + read_file: 无异常、数据不丢
errors = []


def e7_writer(tag: str, n: int):
    try:
        for i in range(n):
            svc.update_cells("t3", [{"sheet": "Sheet", "row": i, "col": 0, "value": f"{tag}-{i}"}])
    except Exception as e:
        errors.append(f"writer: {e}")


svc.create_file("t3")
threads = [threading.Thread(target=e7_writer, args=(f"w{k}", 30)) for k in range(4)]
read_err = []


def e7_reader():
    try:
        for _ in range(20):
            svc.read_file("t3")
            time.sleep(0.01)
    except Exception as e:
        read_err.append(str(e))


threads.append(threading.Thread(target=e7_reader))
for t in threads:
    t.start()
for t in threads:
    t.join(timeout=15)
svc.flush_pending()
data = svc.read_file("t3")
rows = data["sheets"][0]["data"]
check("E7 并发写读零异常", not errors and not read_err, f"err={errors[:2]}{read_err[:2]}")
# 4线程写同30行: 每格最后写赢, 最终30行全部有值(无丢行)
nonempty = [r for r in rows if r and r[0]]
check("E7b 并发写数据完整(30行全落盘无丢失)", len(nonempty) == 30, f"got {len(nonempty)}")

# E8 高频落盘(原子写) + 并发读: 零 BadZipFile
svc.create_file("t4")
stop = threading.Event()
read_stats = {"ok": 0, "err": 0}


def e8_writer():
    n = 0
    while not stop.is_set():
        n += 1
        updates = [{"sheet": "Sheet", "row": r, "col": 0, "value": f"{n}-{r}"} for r in range(100)]
        svc.update_cells("t4", updates)
        time.sleep(0.05)


def e8_reader():
    while not stop.is_set():
        try:
            svc.read_file("t4")
            read_stats["ok"] += 1
        except Exception as e:
            read_stats["err"] += 1
            errors.append(f"e8reader: {type(e).__name__}")
        time.sleep(0.005)


ths = [threading.Thread(target=e8_writer, daemon=True), threading.Thread(target=e8_reader, daemon=True)]
for t in ths:
    t.start()
time.sleep(3)
stop.set()
for t in ths:
    t.join(timeout=10)
check(
    f"E8 原子写下并发读{read_stats['ok']}次零失败(修复前读损坏zip)",
    read_stats["err"] == 0 and read_stats["ok"] > 10,
    f"stats={read_stats}",
)
leftover = [f for f in os.listdir(tmpdir) if f.endswith(".tmp")]
check("E8b 无残留tmp", not leftover, f"got {leftover}")

# ============================================================
# F. AsyncFileWatcher
# ============================================================
print("\n========== F. AsyncFileWatcher ==========")


async def run_watcher_tests():
    f = os.path.join(tmpdir, "watch.xlsx")
    from openpyxl import Workbook

    wb = Workbook()
    wb.save(f)
    wb.close()

    watcher = AsyncFileWatcher(f)
    events = []
    done = asyncio.Event()

    async def consume():
        async for ev in watcher.start():
            events.append(ev)
            if ev["type"] == "file_changed" and len(events) >= 1:
                done.set()

    task = asyncio.create_task(consume())
    await asyncio.sleep(0.5)  # 等 watcher 就绪

    # F1: 修改文件 → watchdog线程回调 → call_soon_threadsafe → 防抖0.5s → file_changed
    wb = load_workbook(f)
    wb.active.cell(row=1, column=1, value="changed")
    wb.save(f)
    wb.close()
    try:
        await asyncio.wait_for(done.wait(), timeout=5)
        check("F1 修改触发 file_changed(线程安全调度)", True)
    except TimeoutError:
        check("F1 修改触发 file_changed(线程安全调度)", False, f"events={events}")

    # F2: pause_watching 期间修改应被忽略
    watcher.pause_watching(duration=1.5)
    n_before = len(events)
    wb = load_workbook(f)
    wb.active.cell(row=2, column=1, value="paused")
    wb.save(f)
    wb.close()
    await asyncio.sleep(1.2)  # pause 窗口内(1.5s)
    got_during_pause = any(e["type"] == "file_changed" for e in events[n_before:])
    check("F2 pause期间修改被忽略", not got_during_pause, f"events={events[n_before:]}")
    await asyncio.sleep(1.5)  # 等 pause 结束后的下一次事件(可能无)

    # F3: 删除文件 → file_deleted 立即
    os.remove(f)
    try:
        for _ in range(50):
            await asyncio.sleep(0.1)
            if any(e["type"] == "file_deleted" for e in events):
                break
        check("F3 删除触发 file_deleted", any(e["type"] == "file_deleted" for e in events), f"events={events}")
    finally:
        watcher.stop()
        task.cancel()
        try:
            await task
        except (asyncio.CancelledError, GeneratorExit):
            pass


asyncio.run(run_watcher_tests())

# 清理
shutil.rmtree(tmpdir, ignore_errors=True)

print(f"\n========== 结果: {PASS} 通过, {FAIL} 失败 ==========")
sys.exit(1 if FAIL else 0)
