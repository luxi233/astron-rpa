import os
import threading
import time
import uuid
from collections.abc import Generator
from typing import Any

from astronverse.scheduler.logger import logger
from openpyxl import Workbook, load_workbook

# 单元格更新合并防抖窗口: 快速连续编辑(如前端逐格输入)合并为一次落盘
_UPDATE_DEBOUNCE_SECONDS = 0.2

# 并发读写容错: 执行器进程原子替换文件的瞬间(或异常时序下)读到瞬时不可用文件时,
# 短间隔重试, 避免前端拉取(fetchDataTable)失败导致表格显示旧数据
_READ_RETRY_TIMES = 3
_READ_RETRY_INTERVAL = 0.1

# Windows 下 os.replace 要求目标文件无其他进程持有句柄(Python 的 open 不带 FILE_SHARE_DELETE):
# 执行器写盘与前端重拉读/执行器侧写并发时 replace 可能报 PermissionError,
# 短间隔重试穿透瞬时锁(总窗口 ~2s), 耗尽后抛出原始异常
_REPLACE_RETRY_TIMES = 10
_REPLACE_RETRY_INTERVAL = 0.2


def _atomic_save(workbook, file_path: str) -> None:
    """原子保存工作簿: 先写同目录临时文件再 os.replace 替换。

    非原子直接覆盖写时, 执行器/前端并发 load_workbook 会读到写一半的损坏
    zip(BadZipFile), 导致数据表格显示不稳定。临时文件带 pid+uuid 防多进程冲突。
    """
    tmp_path = "{}.{}.{}.tmp".format(file_path, os.getpid(), uuid.uuid4().hex[:8])
    try:
        workbook.save(tmp_path)
        last_err = None
        for _ in range(_REPLACE_RETRY_TIMES):
            try:
                os.replace(tmp_path, file_path)
                break
            except PermissionError as e:
                last_err = e
                time.sleep(_REPLACE_RETRY_INTERVAL)
        else:
            raise last_err
    except Exception:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except OSError:
            pass
        raise


def _load_workbook_with_retry(file_path: str, **kwargs):
    """带短重试的 load_workbook: 读到瞬时不可用文件(极端时序/平台差异)时重试,
    重试耗尽仍失败才抛出。配合写侧原子替换, 正常情况下首次即成功。"""
    for attempt in range(_READ_RETRY_TIMES):
        try:
            return load_workbook(file_path, **kwargs)
        except Exception:
            if attempt == _READ_RETRY_TIMES - 1:
                raise
            time.sleep(_READ_RETRY_INTERVAL)


class ExcelService:
    """Excel 文件读写服务"""

    # 合并防抖状态(类级共享: get_excel_service 每次请求新建实例, 状态必须跨实例存活)
    _pending_updates: dict[str, list[dict]] = {}
    _pending_lock = threading.Lock()
    _flush_timer: threading.Timer | None = None
    # flush 重入防护: 读/写接口的 flush_pending 与防抖定时器回调 _flush_pending 可能并发,
    # 两者都在锁外应用更新, 若无防护同一批更新可能被双取并发 apply(读-改-写交叉导致丢写)。
    # 后到者等待进行中的 flush 完成后再处理(保证读接口"先落盘再读"的一致性语义)。
    # Condition 复用 _pending_lock: 队列入队(update_cells)与出队(flush)必须同一把锁
    _flush_cond = threading.Condition(_pending_lock)
    _flushing = False

    def __init__(self, resource_dir: str):
        """
        初始化 Excel 服务

        Args:
            resource_dir: 工程资源目录
        """
        self.resource_dir = resource_dir

    def get_file_path(self, filename: str) -> str:
        """
        获取 Excel 文件的完整路径

        Args:
            filename: 文件名（不含扩展名）

        Returns:
            完整的文件路径
        """
        if not filename.endswith(".xlsx"):
            filename = f"{filename}.xlsx"
        return os.path.join(self.resource_dir, filename)

    def file_exists(self, filename: str) -> bool:
        """
        检查文件是否存在

        Args:
            filename: 文件名

        Returns:
            文件是否存在
        """
        file_path = self.get_file_path(filename)
        return os.path.exists(file_path)

    def create_file(self, filename: str) -> str:
        """
        创建空白 Excel 文件

        Args:
            filename: 文件名

        Returns:
            创建的文件路径
        """
        file_path = self.get_file_path(filename)

        # 确保目录存在
        dir_path = os.path.dirname(file_path)
        if dir_path and not os.path.exists(dir_path):
            os.makedirs(dir_path, exist_ok=True)

        # 创建空白工作簿
        wb = Workbook()
        _atomic_save(wb, file_path)
        wb.close()

        logger.info(f"Created Excel file: {file_path}")
        return file_path

    def read_file_stream(self, filename: str) -> Generator[dict]:
        """
        流式读取 Excel 文件，逐行返回数据

        Args:
            filename: 文件名

        Yields:
            每行数据的字典，格式为 {"sheet": str, "row": int, "data": list}
        """
        # 先落盘待保存的单元格更新, 保证读到最新内容
        self.flush_pending()
        file_path = self.get_file_path(filename)

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        wb = load_workbook(file_path, read_only=True, data_only=True)

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # 发送 sheet 开始事件
                yield {
                    "type": "sheet_start",
                    "sheet": sheet_name,
                    "max_row": ws.max_row or 0,
                    "max_column": ws.max_column or 0,
                }

                # 逐行读取数据
                row_num = 0
                for row in ws.iter_rows(values_only=True):
                    row_num += 1
                    # 将单元格值转换为可序列化的格式
                    row_data = [self._serialize_cell_value(cell) for cell in row]
                    yield {
                        "type": "row",
                        "sheet": sheet_name,
                        "row": row_num,
                        "data": row_data,
                    }

                # 发送 sheet 结束事件
                yield {
                    "type": "sheet_end",
                    "sheet": sheet_name,
                }

        finally:
            wb.close()

        # 发送完成事件
        yield {
            "type": "complete",
            "filename": filename,
        }

    def read_file(self, filename: str) -> dict:
        """
        一次性读取整个 Excel 文件

        Args:
            filename: 文件名

        Returns:
            包含所有数据的字典
        """
        # 先落盘待保存的单元格更新, 保证读到最新内容
        self.flush_pending()
        file_path = self.get_file_path(filename)

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        wb = _load_workbook_with_retry(file_path, read_only=True, data_only=False)

        try:
            result = {
                "filename": filename,
                "sheets": [],
                "active_sheet": wb.active.title if wb.active else None,
            }

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_data = {
                    "name": sheet_name,
                    "max_row": ws.max_row or 0,
                    "max_column": ws.max_column or 0,
                    "data": [],
                }

                for row in ws.iter_rows(values_only=True):
                    row_data = [self._serialize_cell_value(cell) for cell in row]
                    sheet_data["data"].append(row_data)

                # 裁剪尾部全空行/列: 文件 dimension 可能虚高(历史版本样式事件误写 null 撑大),
                # 按真实内容收敛 max_row/max_column, 避免前端展示大量空行
                while sheet_data["data"]:
                    if any(cell is not None and cell != "" for cell in sheet_data["data"][-1]):
                        break
                    sheet_data["data"].pop()
                max_col = 0
                for row_data in sheet_data["data"]:
                    for c in range(len(row_data) - 1, -1, -1):
                        if row_data[c] is not None and row_data[c] != "":
                            max_col = max(max_col, c + 1)
                            break
                for row_data in sheet_data["data"]:
                    del row_data[max_col:]
                sheet_data["max_row"] = len(sheet_data["data"])
                sheet_data["max_column"] = max_col

                result["sheets"].append(sheet_data)

            return result

        finally:
            wb.close()

    def write_file(self, filename: str, data: dict) -> None:
        """
        写入数据到 Excel 文件

        Args:
            filename: 文件名
            data: 要写入的数据，格式为 {"sheets": [{"name": str, "data": list[list]}]}
        """
        # 先落盘待保存的单元格更新, 避免全量覆盖后又被旧pending回写
        self.flush_pending()
        file_path = self.get_file_path(filename)

        # 确保目录存在
        dir_path = os.path.dirname(file_path)
        if dir_path and not os.path.exists(dir_path):
            os.makedirs(dir_path, exist_ok=True)

        wb = Workbook()

        # 删除默认创建的 sheet
        if wb.active:
            wb.remove(wb.active)

        sheets = data.get("sheets", [])
        if not sheets:
            # 如果没有数据，至少创建一个空 sheet
            wb.create_sheet("Sheet1")
        else:
            for sheet_info in sheets:
                sheet_name = sheet_info.get("name", "Sheet1")
                ws = wb.create_sheet(sheet_name)

                sheet_data = sheet_info.get("data", [])
                for row_idx, row_data in enumerate(sheet_data, start=1):
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # 设置活动 sheet
        active_sheet = data.get("active_sheet")
        if active_sheet and active_sheet in wb.sheetnames:
            wb.active = wb[active_sheet]
        elif wb.sheetnames:
            wb.active = wb[wb.sheetnames[0]]

        _atomic_save(wb, file_path)
        wb.close()

        logger.info(f"Saved Excel file: {file_path}")

    def update_cells(self, filename: str, updates: list[dict]) -> None:
        """
        更新指定单元格的值(带合并防抖: 入队后由定时器合并落盘, 读接口自动先flush保证一致)

        Args:
            filename: 文件名
            updates: 更新列表，每项格式为 {"sheet": str, "row": int, "col": int, "value": any}
        """
        file_path = self.get_file_path(filename)

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        with ExcelService._pending_lock:
            ExcelService._pending_updates.setdefault(file_path, []).extend(updates)
            # 已有定时器在等待则复用(持续编辑时滚动延迟), 否则起一个新定时器
            if ExcelService._flush_timer is None:
                ExcelService._flush_timer = threading.Timer(_UPDATE_DEBOUNCE_SECONDS, ExcelService._flush_pending)
                ExcelService._flush_timer.daemon = True
                ExcelService._flush_timer.start()

    @classmethod
    def flush_pending(cls) -> None:
        """立即落盘全部待保存的单元格更新(读/写/删文件前调用保证一致性)"""
        with cls._flush_cond:
            while cls._flushing:
                # 已有 flush 在进行, 等其完成后再检查队列(避免读到未落盘数据)
                cls._flush_cond.wait()
            cls._flushing = True
            items = list(cls._pending_updates.items())
            cls._pending_updates.clear()
            if cls._flush_timer is not None:
                cls._flush_timer.cancel()
                cls._flush_timer = None
        try:
            for file_path, updates in items:
                try:
                    cls._apply_updates(file_path, updates)
                except Exception:
                    # 单个文件失败(如已被删除)不影响其他文件, 丢弃该批待更新
                    logger.exception(f"Flush pending cell updates failed: {file_path}")
        finally:
            with cls._flush_cond:
                cls._flushing = False
                cls._flush_cond.notify_all()

    @classmethod
    def _flush_pending(cls) -> None:
        """定时器回调: 取出并应用全部待更新(防抖静默期到达)"""
        with cls._flush_cond:
            while cls._flushing:
                cls._flush_cond.wait()
            cls._flushing = True
            items = list(cls._pending_updates.items())
            cls._pending_updates.clear()
            cls._flush_timer = None
        try:
            for file_path, updates in items:
                try:
                    cls._apply_updates(file_path, updates)
                except Exception:
                    logger.exception(f"Flush pending cell updates failed: {file_path}")
        finally:
            with cls._flush_cond:
                cls._flushing = False
                cls._flush_cond.notify_all()

    @staticmethod
    def _apply_updates(file_path: str, updates: list[dict]) -> None:
        """将一批单元格更新写入文件(原 update_cells 主体)"""
        wb = load_workbook(file_path)

        try:
            for update in updates:
                sheet_name = update.get("sheet")
                row = update.get("row") + 1
                col = update.get("col") + 1
                value = update.get("value")

                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(sheet_name)
                else:
                    ws = wb[sheet_name]

                # 防御: 空值写入已用区域之外会撑大 max_row/max_column 且无意义, 跳过
                # (前端样式类事件误传整列 null 时避免污染文件)
                if value is None and (row > ws.max_row or col > ws.max_column):
                    continue

                if value is None:
                    # ws.cell(value=None) 不会清空原值(None 表示"不设置"), 需显式赋值清空
                    ws.cell(row=row, column=col).value = None
                else:
                    ws.cell(row=row, column=col, value=value)

            _atomic_save(wb, file_path)
            logger.info(f"Updated {len(updates)} cells in: {file_path}")

        finally:
            wb.close()

    def delete_file(self, filename: str) -> bool:
        """
        删除 Excel 文件

        Args:
            filename: 文件名

        Returns:
            是否删除成功
        """
        # 先落盘待保存的单元格更新(文件即将删除, 清空pending避免定时器写已删除文件)
        self.flush_pending()
        file_path = self.get_file_path(filename)

        if os.path.exists(file_path):
            os.remove(file_path)
            logger.info(f"Deleted Excel file: {file_path}")
            return True

        return False

    @staticmethod
    def _serialize_cell_value(value: Any) -> Any:
        """
        将单元格值转换为可 JSON 序列化的格式

        Args:
            value: 单元格原始值

        Returns:
            序列化后的值
        """
        if value is None:
            return None
        elif isinstance(value, (str, int, float, bool)):
            return value
        else:
            # 其他类型转换为字符串
            return str(value)
